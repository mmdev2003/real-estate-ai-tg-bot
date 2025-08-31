import io
import os
import datetime as DT
from uuid import uuid4

import numpy_financial as np_f
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.cell import get_column_letter

from opentelemetry.trace import Status, StatusCode, SpanKind

from internal import interface, common, model


class NeedRepairs:
    need_repairs = 0
    dont_need_repairs = 1
    need_partial_repairs = 2


class EstateCalculator(interface.IEstateCalculator):

    def __init__(
            self,
            tel: interface.ITelemetry,
            metro_repo: interface.IMetroRepo,
            income_tax_share: float
    ):
        self.tracer = tel.tracer()
        self.metro_repo = metro_repo
        self.income_tax_share = income_tax_share

    async def calc_finance_model_finished_office(
            self,
            square: float,
            price_per_meter: float,
            need_repairs: int,
            metro_station_name: str,
            estate_category: str,
            distance_to_metro: float,
            nds_rate: int,
            create_xlsx: bool = False,
    ) -> dict | io.BytesIO:
        """
        Generate a irr calculations on the profitability of sale real estate.
        And generate excell sheet with calculations.

        Args:
        square (float) - Total area in square meter
        price_per_meter (float) - Price in rubles per square meter
        distance_to_metro (float) - The path from real estate to the subway in meter
        metro_station_name (str) - Name of the metro station
        estate_category (str) - Real estate category (A or B)
        number_of_quartals (int) - The number of quartals to calculate
        sale_quartal (str) - quarter of the year when the property is for sale in format (1Q2024)
        transaction_dict (dict) - information about the frequency and percentage of transactions
        average_value_real_estate (int) - Average specific value of real estate
        minimal_ipc (float) - minimal ipc
        create_report (bool) - Do I need create report in excel file
        rental_holidays (int) - Rental holidays (office renovation) in month
        price_of_finishing (float) - The price of finishing the room in rubles per square meter
        capitalization_rate (float) - Capitalization rate Kp in %
        rent_indexing (float) - Annual rental indexation in %
        need_repairs (Enum int) - 0 - need repairs< 1 - dont need, 2 - need 80% of repairs

        Returns:
        exls_sheet:  Returns the excell sheet with calculations
        irr (float): Returns the rental price of real estate.
        """
        with self.tracer.start_as_current_span(
                "EstateCalculator.calc_finance_model_finished_office",
                kind=SpanKind.INTERNAL,
                attributes={
                    "square": square,
                    "price_per_meter": price_per_meter,
                    "need_repairs": need_repairs,
                    "metro_station_name": metro_station_name,
                    "estate_category": estate_category,
                    "distance_to_metro": distance_to_metro,
                    "nds_rate": nds_rate,
                    "create_xlsx": create_xlsx
                }
        ) as span:
            try:
                transaction_dict: dict = {"1Q2025": 100}
                minimal_ipc: float = 6.0
                number_of_quartals: int = 23
                rental_holidays: int = 6
                price_of_finishing: float = 70_000.0
                capitalization_rate: float = 8.5
                rent_indexing: float = 7.0

                if NeedRepairs.dont_need_repairs == need_repairs:
                    price_of_finishing = 0
                elif NeedRepairs.need_partial_repairs == need_repairs:
                    price_of_finishing *= 0.8

                # styles
                ff_3 = PatternFill("solid", fgColor="00FFFF99")
                ff_4 = PatternFill("solid", fgColor="00C0C0C0")
                thins = Side(border_style="thin", color="00808080")
                bordr_all = Border(left=thins, right=thins, bottom=thins, top=thins)
                bordr_top = Border(top=thins)
                bordr_bottom = Border(bottom=thins)
                bordr_left = Border(left=thins)
                bordr_right = Border(right=thins)
                bordr_right_top = Border(right=thins, top=thins)
                bordr_left_top = Border(left=thins, top=thins)
                bordr_right_bottom = Border(right=thins, bottom=thins)
                bordr_left_bottom = Border(left=thins, bottom=thins)

                EPA = self.EPA_validation(self.first_quartal(), self.first_quartal(), transaction_dict)
                project_readiness = EPA

                wb = Workbook()

                exls_sheet = wb.active
                exls_sheet.title = "Готовые"

                exls_sheet.column_dimensions["B"].width = 70
                exls_sheet.column_dimensions["C"].width = 30
                exls_sheet["B1"] = "Финансовая модель оценки инвестиционной привлекательности готового офиса"
                exls_sheet.merge_cells("B1:C1")

                exls_sheet["B2"] = "Дата сделки"
                exls_sheet["C2"] = self.first_quartal()

                exls_sheet["B3"] = "Строительная готовность"
                exls_sheet["C3"] = project_readiness

                exls_sheet["B5"] = "Данные по выбранному помещению"

                exls_sheet["B6"] = "Лот"
                exls_sheet["C6"] = "-------"

                exls_sheet["B7"] = "Общая площадь"
                exls_sheet["C7"] = self.division_into_categories(square)
                exls_sheet["D7"] = "кв.м"

                exls_sheet["B8"] = "Цена"
                exls_sheet["C8"] = self.division_into_categories(price_per_meter)
                exls_sheet["D8"] = "руб./ кв. м."

                exls_sheet["B9"] = "Цена Лота (Total)"
                exls_sheet["C9"] = self.division_into_categories(square * price_per_meter)
                exls_sheet["D9"] = "руб."

                price_rva = await self.calculate_price(
                    square=square,
                    distance_to_metro=distance_to_metro,
                    metro_station_name=metro_station_name,
                    estate_category=estate_category,
                    strategy="price",
                )

                rental_rate = await self.calculate_price(
                    square=square,
                    distance_to_metro=distance_to_metro,
                    metro_station_name=metro_station_name,
                    estate_category=estate_category,
                    strategy="rent",
                )

                average_value_real_estate = await self.get_cadastral_value(metro_station_name=metro_station_name)

                Q_years_list, years_list = self.create_datas_list(EPA, project_readiness, number_of_quartals)
                sale_quartal = Q_years_list[Q_years_list.index(project_readiness) + 1]
                self.validation_transactions_dict(transaction_dict)
                transaction_dict = self.create_transactions_dict(transaction_dict)

                exls_sheet["B11"] = "Условия рассрочки"

                exls_sheet["B12"] = "Размер"

                exls_sheet["B13"] = "Квартал"

                exls_sheet["B14"] = "Сумма траншей"
                ft_3 = Font(color="00FF0000")
                for r, quartal in enumerate(Q_years_list):
                    exls_sheet.cell(row=13, column=r + 4).value = quartal
                    if quartal in transaction_dict.keys():
                        exls_sheet.cell(row=12, column=r + 4).value = transaction_dict[quartal]
                        exls_sheet.cell(row=14, column=r + 4).value = self.division_into_categories(
                            -transaction_dict[quartal] * square * price_per_meter
                        )
                        exls_sheet.cell(row=14, column=r + 4).font = ft_3
                exls_sheet["C14"] = self.division_into_categories(
                    -sum([transaction_dict[quartal] * square * price_per_meter for quartal in transaction_dict.keys()])
                )
                exls_sheet["C14"].font = ft_3
                exls_sheet["C12"] = "100%"

                exls_sheet["B16"] = "Базовые предпосылки к расчету"

                exls_sheet["B17"] = "Перепродажа"

                rental_holidays = rental_holidays // 3
                if int(project_readiness[0]) + rental_holidays <= 4:
                    rental_date = str(int(project_readiness[0]) + rental_holidays)
                else:
                    rental_date = str(int(project_readiness[0]) + rental_holidays - 4)

                if int(project_readiness[0]) + 1 <= 4:
                    renovation_quartal = str(int(project_readiness[0]) + 1) + "Q" + str(int(project_readiness[2:]))
                else:
                    renovation_quartal = str(int(project_readiness[0]) + 1 - 4) + "Q" + str(
                        int(project_readiness[2:]) + 1)

                if int(project_readiness[0]) + 1 <= 4:
                    rental_flow_quartal = str(int(project_readiness[0]) + 1) + "Q" + str(int(project_readiness[2:]))
                else:
                    rental_flow_quartal = str(int(project_readiness[0]) + 1 - 4) + "Q" + str(
                        int(project_readiness[2:]) + 1)

                exls_sheet["B18"] = "Получения арендного потока"
                exls_sheet["C18"] = str(rental_date) + "Q"
                exls_sheet["C19"] = rental_flow_quartal

                exls_sheet["B21"] = "Ставка аренды по объекту на дату покупки"
                exls_sheet["C21"] = self.division_into_categories(rental_rate)
                exls_sheet["D21"] = "руб./кв.м./год"

                exls_sheet["B22"] = "МАП"
                m_a_p = rental_rate / 12 * square
                exls_sheet["C22"] = self.division_into_categories(m_a_p)
                m_a_p = rental_rate / 4 * square
                exls_sheet["D22"] = "руб. в месяц"

                exls_sheet["B23"] = "Цена отделки помещения (Fit-Out)"
                exls_sheet["C23"] = self.division_into_categories(price_of_finishing)
                exls_sheet["D23"] = "руб./кв.м."

                exls_sheet["B24"] = "Арендные каникулы (ремонт офиса)"
                exls_sheet["C24"] = rental_holidays * 3 / 2
                exls_sheet["D24"] = "мес"

                capitalization_rate /= 100
                exls_sheet["B25"] = "Ставка капитализации Kp"
                exls_sheet["C25"] = str(round(capitalization_rate * 100, 1)) + "%"

                exls_sheet["B26"] = "Дата расчета Terminal value"
                exls_sheet["C26"] = Q_years_list[-1]

                rent_indexing /= 100
                exls_sheet["B27"] = "Ежегодная индексация аренды, %"
                exls_sheet["C27"] = str(round(rent_indexing * 100, 1)) + "%"

                exls_sheet["B28"] = "Средняя удельная КС недвижимости"
                exls_sheet["C28"] = self.division_into_categories(average_value_real_estate)
                exls_sheet["D28"] = "руб./кв.м."

                exls_sheet["B29"] = "НДС на УСН"
                exls_sheet["C29"] = str(nds_rate) + "%"

                exls_sheet["B30"] = "Дата оплаты первого налога на недвижимость"
                exls_sheet["C30"] = Q_years_list[Q_years_list.index(project_readiness) + 4]
                # exls_sheet["C30"] = Q_years_list[0]

                list_f_border_0 = ["B2", "B6", "B12", "C12", "B17", "B49", "B65"]
                for cell in list_f_border_0:
                    exls_sheet[cell].border = bordr_left_top

                list_f_border_1 = ["B3", "B9", "B14", "C14", "B30", "B58", "B74"]
                for cell in list_f_border_1:
                    exls_sheet[cell].border = bordr_left_bottom

                list_f_border_2 = ["C6", "D6", "D17", "C65"]
                for cell in list_f_border_2:
                    exls_sheet[cell].border = bordr_right_top

                list_f_border_3 = ["C9", "D9", "D30", "C74"]
                for cell in list_f_border_3:
                    exls_sheet[cell].border = bordr_right_bottom

                for r in range(49, 59):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(66, 75):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(61, 63):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(77, 79):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(7, 9):
                    for c in range(1, 5):
                        if c != 2:
                            exls_sheet.cell(row=r, column=c).border = bordr_right

                for r in range(13, 14):
                    for c in range(1, 3):
                        exls_sheet.cell(row=r, column=c).border = bordr_right

                for r in range(18, 30):
                    for c in range(1, 2):
                        exls_sheet.cell(row=r, column=c).border = bordr_right

                for r in range(18, 30):
                    for c in range(4, 5):
                        exls_sheet.cell(row=r, column=c).border = bordr_right

                exls_sheet["C17"].border = bordr_top
                exls_sheet["C30"].border = bordr_bottom

                exls_sheet["C2"].border = bordr_all
                exls_sheet["C3"].border = bordr_all

                lot_price = square * price_per_meter
                rent_qurtals_dict = {}
                rent_flow = 0
                for quartal in Q_years_list:
                    rent_qurtals_dict[quartal] = {
                        "auxiliary_calculation": 0.0,
                        "buying_property": 0.0,
                        "rent_flow": 0.0,
                        "income_tax": 0.0,
                        "renovation": 0.0,
                        "terminal_value": 0.0,
                        "property_tax": 0.0,
                        "money_flow": 0.0,
                    }
                    if quartal[0] == rental_date:
                        rent_flow += rent_flow * rent_indexing
                        rent_qurtals_dict[quartal]["auxiliary_calculation"] = rent_indexing
                    if quartal == renovation_quartal:
                        rent_qurtals_dict[quartal]["renovation"] = -square * price_of_finishing
                    if quartal == Q_years_list[Q_years_list.index(project_readiness) + rental_holidays - 1]:
                        rent_flow = m_a_p
                    rent_qurtals_dict[quartal]["rent_flow"] = rent_flow
                    if quartal in transaction_dict.keys():
                        rent_qurtals_dict[quartal]["buying_property"] = -lot_price * transaction_dict[quartal]
                    if Q_years_list.index(quartal) == len(Q_years_list) - 1:
                        rent_qurtals_dict[quartal]["terminal_value"] = (
                                rent_qurtals_dict[quartal]["rent_flow"] * 4 / capitalization_rate
                        )
                    if Q_years_list.index(quartal) >= Q_years_list.index(project_readiness) + 4:
                        rent_qurtals_dict[quartal]["property_tax"] = -average_value_real_estate * square / 4 * 0.015

                    rent_qurtals_dict[quartal]["income_tax"] = (
                            -(rent_qurtals_dict[quartal]["rent_flow"] + rent_qurtals_dict[quartal]["terminal_value"])
                            * (self.income_tax_share + (nds_rate / 100))
                    )
                    rent_qurtals_dict[quartal]["money_flow"] = (
                            rent_qurtals_dict[quartal]["buying_property"]
                            + rent_qurtals_dict[quartal]["rent_flow"]
                            + rent_qurtals_dict[quartal]["income_tax"]
                            + rent_qurtals_dict[quartal]["renovation"]
                            + rent_qurtals_dict[quartal]["terminal_value"]
                            + rent_qurtals_dict[quartal]["property_tax"]
                    )

                rent_irr = (
                                   (1 + np_f.irr(
                                       [rent_qurtals_dict[key]["money_flow"] for key in
                                        rent_qurtals_dict.keys()])) ** 4 - 1
                           ) * 100
                if np.isnan(rent_irr):
                    rent_irr = -1

                ali_1 = Alignment(horizontal="center", vertical="center")
                for r, quartal in enumerate(Q_years_list):
                    col = get_column_letter(r + 4)
                    l_col = get_column_letter(r + 3)

                    exls_sheet.cell(row=66, column=r + 4).value = quartal
                    if str(rental_date) + "Q" in quartal:
                        exls_sheet.cell(row=65, column=r + 4).value = rent_indexing
                    else:
                        exls_sheet.cell(row=65, column=r + 4).value = 0

                    if quartal in transaction_dict.keys():
                        exls_sheet.cell(row=67, column=r + 4).value = str(transaction_dict[quartal] * 100) + "%"
                        exls_sheet.cell(row=68, column=r + 4).value = self.division_into_categories(
                            -transaction_dict[quartal] * square * price_per_meter
                        )
                        exls_sheet.cell(row=68, column=r + 4).font = ft_3
                    if rent_qurtals_dict[quartal]["rent_flow"] < 0:
                        exls_sheet.cell(row=69, column=r + 4).font = ft_3
                    exls_sheet.cell(row=69, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["rent_flow"]
                    )

                    if rent_qurtals_dict[quartal]["renovation"] < 0:
                        exls_sheet.cell(row=71, column=r + 4).font = ft_3
                    exls_sheet.cell(row=71, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["renovation"]
                    )

                    if rent_qurtals_dict[quartal]["terminal_value"] < 0:
                        exls_sheet.cell(row=72, column=r + 4).font = ft_3
                    exls_sheet.cell(row=72, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["terminal_value"]
                    )

                    if rent_qurtals_dict[quartal]["income_tax"] < 0:
                        exls_sheet.cell(row=70, column=r + 4).font = ft_3
                    exls_sheet.cell(row=70, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["income_tax"]
                    )

                    if rent_qurtals_dict[quartal]["property_tax"] < 0:
                        exls_sheet.cell(row=73, column=r + 4).font = ft_3
                    exls_sheet.cell(row=73, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["property_tax"]
                    )

                    if rent_qurtals_dict[quartal]["money_flow"] < 0:
                        exls_sheet.cell(row=74, column=r + 4).font = ft_3
                    exls_sheet.cell(row=74, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["money_flow"]
                    )

                    exls_sheet.cell(row=12, column=r + 4).fill = ff_3
                    exls_sheet.cell(row=13, column=r + 4).fill = ff_3
                    exls_sheet.cell(row=67, column=r + 4).fill = ff_4

                    for i in range(65, 75):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all

                    for i in range(12, 15):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all

                    exls_sheet.column_dimensions[col].width = 15

                for r in range(1, 150):
                    for c in range(3, 150):
                        exls_sheet.cell(row=r, column=c).alignment = ali_1
                year_dict = {}
                for r, year in enumerate(years_list):
                    col = get_column_letter(r + 4)
                    year = str(year)
                    year_dict[year] = {
                        "transactions_percent": sum(
                            [transaction_dict[quartal] for quartal in transaction_dict.keys() if year in quartal]
                        ),
                        "buying_property": sum(
                            [
                                rent_qurtals_dict[quartal]["buying_property"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "rent_flow": sum(
                            [rent_qurtals_dict[quartal]["rent_flow"] for quartal in rent_qurtals_dict.keys() if
                             year in quartal]
                        ),
                        "income_tax": sum(
                            [
                                rent_qurtals_dict[quartal]["income_tax"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "renovation": sum(
                            [
                                rent_qurtals_dict[quartal]["renovation"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "terminal_value": sum(
                            [
                                rent_qurtals_dict[quartal]["terminal_value"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "property_tax": sum(
                            [
                                rent_qurtals_dict[quartal]["property_tax"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "money_flow": sum(
                            [
                                rent_qurtals_dict[quartal]["money_flow"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                    }
                    exls_sheet.cell(row=49, column=r + 4).value = year

                    if year_dict[year]["transactions_percent"] < 0:
                        exls_sheet.cell(row=51, column=r + 4).font = ft_3
                    exls_sheet.cell(row=51, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["transactions_percent"]
                    )

                    if year_dict[year]["buying_property"] < 0:
                        exls_sheet.cell(row=52, column=r + 4).font = ft_3
                    exls_sheet.cell(row=52, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["buying_property"]
                    )

                    if year_dict[year]["rent_flow"] < 0:
                        exls_sheet.cell(row=53, column=r + 4).font = ft_3
                    exls_sheet.cell(row=53, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["rent_flow"])

                    if year_dict[year]["terminal_value"] < 0:
                        exls_sheet.cell(row=54, column=r + 4).font = ft_3
                    exls_sheet.cell(row=54, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["terminal_value"]
                    )

                    if year_dict[year]["renovation"] < 0:
                        exls_sheet.cell(row=55, column=r + 4).font = ft_3
                    exls_sheet.cell(row=55, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["renovation"])

                    if year_dict[year]["income_tax"] < 0:
                        exls_sheet.cell(row=56, column=r + 4).font = ft_3
                    exls_sheet.cell(row=56, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["income_tax"])

                    if year_dict[year]["property_tax"] < 0:
                        exls_sheet.cell(row=57, column=r + 4).font = ft_3
                    exls_sheet.cell(row=57, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["property_tax"])

                    if year_dict[year]["money_flow"] < 0:
                        exls_sheet.cell(row=58, column=r + 4).font = ft_3
                    exls_sheet.cell(row=58, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["money_flow"])

                    exls_sheet.cell(row=51, column=r + 4).fill = ff_4

                    for i in range(49, 59):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all

                exls_sheet["B48"] = "2. Сдача в аренду"
                exls_sheet["B51"] = "Покупка"
                exls_sheet["B52"] = "Покупка"
                exls_sheet["B53"] = "Арендный поток за весь период"
                exls_sheet["B54"] = "Terminal value "
                exls_sheet["B55"] = "Ремонт помещения (Fit-Out)"
                exls_sheet["B56"] = f"Налог на доход УСН 6% + НДС для УСН"
                exls_sheet["B57"] = "Налог на недвижимость"
                exls_sheet["B58"] = "Денежный поток"

                exls_sheet["B61"] = "Прибыль"
                exls_sheet["B62"] = "IRR (доходность за период 6 лет с даты покупки)"

                exls_sheet["B64"] = "Вспомогательный расчет"
                exls_sheet["B67"] = "Покупка"
                exls_sheet["B68"] = "Покупка"
                exls_sheet["B69"] = "Арендный поток за весь период"
                exls_sheet["B70"] = f"Налог на доход УСН 6% + НДС для УСН"
                exls_sheet["B71"] = "Ремонт помещения (Fit-Out)"
                exls_sheet["B72"] = "Terminal Value "
                exls_sheet["B73"] = "Налог на недвижимость"
                exls_sheet["B74"] = "Денежный поток"

                exls_sheet["B77"] = "Прибыль"
                exls_sheet["B78"] = "IRR (доходность за период 6 лет с даты покупки)"

                total_dict = {
                    "buying_property": sum(
                        [rent_qurtals_dict[quartal]["buying_property"] for quartal in rent_qurtals_dict.keys()]
                    ),
                    "rent_flow": sum([rent_qurtals_dict[quartal]["rent_flow"] for quartal in rent_qurtals_dict.keys()]),
                    "income_tax": sum(
                        [rent_qurtals_dict[quartal]["income_tax"] for quartal in rent_qurtals_dict.keys()]),
                    "renovation": sum(
                        [rent_qurtals_dict[quartal]["renovation"] for quartal in rent_qurtals_dict.keys()]),
                    "terminal_value": sum(
                        [rent_qurtals_dict[quartal]["terminal_value"] for quartal in rent_qurtals_dict.keys()]
                    ),
                    "property_tax": sum(
                        [rent_qurtals_dict[quartal]["property_tax"] for quartal in rent_qurtals_dict.keys()]),
                    "money_flow": sum(
                        [rent_qurtals_dict[quartal]["money_flow"] for quartal in rent_qurtals_dict.keys()]),
                }
                exls_sheet["C66"] = "Итого"

                exls_sheet["C67"] = "100%"

                if total_dict["buying_property"] < 0:
                    exls_sheet["C68"].font = ft_3
                    exls_sheet["C52"].font = ft_3
                exls_sheet["C68"] = self.division_into_categories(total_dict["buying_property"])
                exls_sheet["C52"] = self.division_into_categories(total_dict["buying_property"])

                if total_dict["rent_flow"] < 0:
                    exls_sheet["C69"].font = ft_3
                    exls_sheet["C53"].font = ft_3
                exls_sheet["C69"] = self.division_into_categories(total_dict["rent_flow"])
                exls_sheet["C53"] = self.division_into_categories(total_dict["rent_flow"])

                if total_dict["income_tax"] < 0:
                    exls_sheet["C70"].font = ft_3
                    exls_sheet["C56"].font = ft_3
                exls_sheet["C70"] = self.division_into_categories(total_dict["income_tax"])
                exls_sheet["C56"] = self.division_into_categories(total_dict["income_tax"])

                if total_dict["renovation"] < 0:
                    exls_sheet["C71"].font = ft_3
                    exls_sheet["C55"].font = ft_3
                exls_sheet["C71"] = self.division_into_categories(total_dict["renovation"])
                exls_sheet["C55"] = self.division_into_categories(total_dict["renovation"])

                if total_dict["terminal_value"] < 0:
                    exls_sheet["C72"].font = ft_3
                    exls_sheet["C54"].font = ft_3
                exls_sheet["C72"] = self.division_into_categories(total_dict["terminal_value"])
                exls_sheet["C54"] = self.division_into_categories(total_dict["terminal_value"])

                if total_dict["property_tax"] < 0:
                    exls_sheet["C73"].font = ft_3
                    exls_sheet["C57"].font = ft_3
                exls_sheet["C73"] = self.division_into_categories(total_dict["property_tax"])
                exls_sheet["C57"] = self.division_into_categories(total_dict["property_tax"])
                sale_tax = total_dict["property_tax"]
                rent_tax = float(total_dict["property_tax"]) + float(total_dict["income_tax"])

                if total_dict["money_flow"] < 0:
                    exls_sheet["C74"].font = ft_3
                    exls_sheet["C58"].font = ft_3
                    exls_sheet["C77"].font = ft_3
                    exls_sheet["C61"].font = ft_3
                exls_sheet["C74"] = self.division_into_categories(total_dict["money_flow"])
                exls_sheet["C58"] = self.division_into_categories(total_dict["money_flow"])
                exls_sheet["C77"] = self.division_into_categories(total_dict["money_flow"])
                exls_sheet["C61"] = self.division_into_categories(total_dict["money_flow"])
                sale_income = total_dict["money_flow"]
                rent_income = total_dict["money_flow"]
                exls_sheet["C49"] = "Итого"
                exls_sheet["C51"] = "100%"

                exls_sheet["C78"] = str(round(rent_irr, 1)) + "%"

                exls_sheet["C62"] = str(round(rent_irr, 1)) + "%"

                ft_1 = Font(name="Calibri", size=11)
                ft_11 = Font(bold=True, name="Calibri", size=14)
                exls_sheet["C78"].font = ft_11
                exls_sheet["C62"].font = ft_11

                ft_1 = Font(name="Calibri", size=11)
                ff_1 = PatternFill("solid", fgColor="0099CCFF")

                ff_2 = PatternFill("solid", fgColor="0099CCFF")
                ft_3 = Font()
                # применяем стиль к ячейкам
                exls_sheet["B1"].font = ft_1
                exls_sheet["B1"].fill = ff_1
                exls_sheet["B48"].font = ft_1
                exls_sheet["B48"].fill = ff_1

                exls_sheet["C20"].fill = ff_2
                exls_sheet["C21"].fill = ff_2
                exls_sheet["C29"].fill = ff_2
                exls_sheet["B51"].fill = ff_4
                exls_sheet["C51"].fill = ff_4
                exls_sheet["B67"].fill = ff_4
                exls_sheet["C67"].fill = ff_4

                exls_sheet["B61"].fill = ff_4
                exls_sheet["C61"].fill = ff_4
                exls_sheet["B62"].fill = ff_4
                exls_sheet["C62"].fill = ff_4

                exls_sheet["B77"].fill = ff_4
                exls_sheet["C77"].fill = ff_4
                exls_sheet["B78"].fill = ff_4
                exls_sheet["C78"].fill = ff_4

                for row in range(32, 47):
                    exls_sheet.row_dimensions[row].hidden = True

                # for key in qurtals_dict.keys():
                #    print(key, end=' ')
                #    for und_key in qurtals_dict[key]:
                #        print(f'{und_key[0:5]}: {qurtals_dict[key][und_key]}', end='\t')
                #    print()

                if create_xlsx:
                    table_buffer = io.BytesIO()
                    wb.save(table_buffer)
                    table_buffer.seek(0)

                    return table_buffer

                sale_property = 0
                sale_irr = 0
                added_value = 0
                result_dict = {
                    "buying_property": round(float(total_dict["buying_property"]) / 1000000, 1),
                    "sale_property": round(sale_property / 1000000, 1),
                    "sale_tax": round(sale_tax / 1000000, 1),
                    "rent_tax": round(rent_tax / 1000000, 1),
                    "price_of_finishing": round(float(total_dict["renovation"]) / 1000000, 1),
                    "rent_flow": round(float(total_dict["rent_flow"]) / 1000000, 1),
                    "terminal_value": round(float(total_dict["terminal_value"]) / 1000000, 1),
                    "sale_income": round(float(sale_income) / 1000000, 1),
                    "rent_income": round(float(rent_income) / 1000000, 1),
                    "added_value": round(float(added_value) / 1000000, 1),
                    "rent_irr": round(rent_irr, 1),
                    "sale_irr": round(sale_irr, 1),
                }

                span.set_status(Status(StatusCode.OK))
                return result_dict
            except Exception as err:
                span.record_exception(err)
                span.set_status(Status(StatusCode.ERROR, str(err)))
                raise err

    async def calc_finance_model_finished_retail(
            self,
            square: float,
            price_per_meter: float,
            m_a_p: float,
            nds_rate: int,
            need_repairs: int,
            create_xlsx: bool = False,
    ) -> dict | io.BytesIO:
        """
        Generate a irr calculations on the profitability of sale retail.
        And generate excell sheet with calculations.

        Args:
        square (float) - Total area in square meter
        price_per_meter (float) - Price in rubles per square meter
        metro_station_name (str) - Name of the metro station
        rental_rate (float) - Result rent of retail in rubles per square meter
        number_of_quartals (int) - The number of quartals to calculate
        sale_quartal (str) - quarter of the year when the property is for sale in format (1Q2024)
        transaction_dict (dict) - information about the frequency and percentage of transactions
        average_value_real_estate (int) - Average specific value of real estate
        minimal_ipc (float) - minimal ipc
        create_report (bool) - Do I need create report in excel file
        rental_holidays (int) - Rental holidays (office renovation) in month
        price_of_finishing (float) - The price_per_meter of finishing the room in rubles per square meter
        capitalization_rate (float) - Capitalization rate Kp in %
        rent_indexing (float) - Annual rental indexation in %
        need_repairs (Enum int) - 0 - need repairs< 1 - dont need, 2 - need 80% of repairs

        Returns:
        exls_sheet:  Returns the excell sheet with calculations
        irr (float): Returns the rental price_per_meter of real estate.
        """
        with self.tracer.start_as_current_span(
                "EstateCalculator.calc_finance_model_finished_retail",
                kind=SpanKind.INTERNAL,
                attributes={
                    "square": square,
                    "price_per_meter": price_per_meter,
                    "m_a_p": m_a_p,
                    "nds_rate": nds_rate,
                    "need_repairs": need_repairs,
                    "create_xlsx": create_xlsx
                }
        ) as span:
            try:
                transaction_dict: dict = {"2Q2025": 100}
                minimal_ipc: float = 6.0
                number_of_quartals: int = 23
                rental_holidays: int = 6
                price_of_finishing: float = 70_000.0
                capitalization_rate: float = 8.5
                rent_indexing: float = 7.0

                if NeedRepairs.dont_need_repairs == need_repairs:
                    price_of_finishing = 0
                elif NeedRepairs.need_partial_repairs == need_repairs:
                    price_of_finishing *= 0.8

                # styles
                ff_3 = PatternFill("solid", fgColor="00FFFF99")
                ff_4 = PatternFill("solid", fgColor="00C0C0C0")
                thins = Side(border_style="thin", color="00808080")
                bordr_all = Border(left=thins, right=thins, bottom=thins, top=thins)
                bordr_top = Border(top=thins)
                bordr_bottom = Border(bottom=thins)
                bordr_left = Border(left=thins)
                bordr_right = Border(right=thins)
                bordr_right_top = Border(right=thins, top=thins)
                bordr_left_top = Border(left=thins, top=thins)
                bordr_right_bottom = Border(right=thins, bottom=thins)
                bordr_left_bottom = Border(left=thins, bottom=thins)

                EPA = self.EPA_validation(self.first_quartal(), self.first_quartal(), transaction_dict)
                project_readiness = EPA

                os.makedirs("/src/reports", exist_ok=True)
                report_file_name = f"/src/reports/{uuid4()}.xlsx"
                wb = Workbook()

                exls_sheet = wb.active
                exls_sheet.title = "Готовые"

                exls_sheet.column_dimensions["B"].width = 40
                exls_sheet.column_dimensions["C"].width = 30
                exls_sheet["B1"] = "Универсальный калькулятор для расчета финмодели"
                exls_sheet.merge_cells("B1:C1")

                exls_sheet["B2"] = "Дата сделки"
                exls_sheet["C2"] = self.first_quartal()

                exls_sheet["B3"] = "Строительная готовность"
                exls_sheet["C3"] = project_readiness

                exls_sheet["B5"] = "Данные по выбранному помещению"

                exls_sheet["B6"] = "Лот"
                exls_sheet["C6"] = "-------"

                exls_sheet["B7"] = "Общая площадь"
                exls_sheet["C7"] = self.division_into_categories(square)
                exls_sheet["D7"] = "кв.м"

                exls_sheet["B8"] = "Цена"
                exls_sheet["C8"] = self.division_into_categories(price_per_meter)
                exls_sheet["D8"] = "руб./ кв. м."

                exls_sheet["B9"] = "Цена Лота (Total)"
                exls_sheet["C9"] = self.division_into_categories(square * price_per_meter)
                exls_sheet["D9"] = "руб."

                average_value_real_estate = 120000
                # average_value_real_estate = self.get_cadastral_value(metro_station_name=metro_station_name)

                Q_years_list, years_list = self.create_datas_list(EPA, project_readiness, number_of_quartals)

                self.validation_transactions_dict(transaction_dict)
                transaction_dict = self.create_transactions_dict(transaction_dict)

                exls_sheet["B11"] = "Условия рассрочки"

                exls_sheet["B12"] = "Размер"

                exls_sheet["B13"] = "Квартал"

                exls_sheet["B14"] = "Сумма траншей"
                ft_3 = Font(color="00FF0000")
                for r, quartal in enumerate(Q_years_list):
                    exls_sheet.cell(row=13, column=r + 4).value = quartal
                    if quartal in transaction_dict.keys():
                        exls_sheet.cell(row=12, column=r + 4).value = str(transaction_dict[quartal] * 100) + "%"
                        exls_sheet.cell(row=14, column=r + 4).value = self.division_into_categories(
                            -transaction_dict[quartal] * square * price_per_meter
                        )
                        exls_sheet.cell(row=14, column=r + 4).font = ft_3
                exls_sheet["C14"] = self.division_into_categories(
                    -sum([transaction_dict[quartal] * square * price_per_meter for quartal in transaction_dict.keys()])
                )
                exls_sheet["C14"].font = ft_3
                exls_sheet["C12"] = "100%"

                exls_sheet["B16"] = "Базовые предпосылки к расчету"

                exls_sheet["B17"] = "Перепродажа"

                rental_holidays = rental_holidays // 3
                if int(project_readiness[0]) + rental_holidays <= 4:
                    rental_date = str(int(project_readiness[0]) + rental_holidays)
                else:
                    rental_date = str(int(project_readiness[0]) + rental_holidays - 4)

                if int(project_readiness[0]) + 1 <= 4:
                    renovation_quartal = str(int(project_readiness[0]) + 1) + "Q" + str(int(project_readiness[2:]))
                else:
                    renovation_quartal = str(int(project_readiness[0]) + 1 - 4) + "Q" + str(
                        int(project_readiness[2:]) + 1)

                if int(project_readiness[0]) + 1 <= 4:
                    rental_flow_quartal = str(int(project_readiness[0]) + 1) + "Q" + str(int(project_readiness[2:]))
                else:
                    rental_flow_quartal = str(int(project_readiness[0]) + 1 - 4) + "Q" + str(
                        int(project_readiness[2:]) + 1)

                exls_sheet["B18"] = "Получения арендного потока"
                exls_sheet["C18"] = str(rental_date) + "Q"
                exls_sheet["C19"] = rental_flow_quartal

                exls_sheet["B20"] = "Цена на РВЭ"
                exls_sheet["C20"] = "-"
                exls_sheet["D20"] = "руб./кв. м."

                rental_rate = m_a_p * 12 / square
                exls_sheet["B21"] = "Ставка аренды"
                exls_sheet["C21"] = self.division_into_categories(rental_rate)
                exls_sheet["D21"] = "руб./кв.м./год"

                exls_sheet["B22"] = "МАП"
                exls_sheet["C22"] = self.division_into_categories(m_a_p)
                exls_sheet["D22"] = "руб. в месяц"

                exls_sheet["B23"] = "Цена отделки помещения (Fit-Out)"
                exls_sheet["C23"] = self.division_into_categories(price_of_finishing)
                exls_sheet["D23"] = "руб./кв.м."

                exls_sheet["B24"] = "Арендные каникулы (ремонт офиса)"
                exls_sheet["C24"] = rental_holidays * 3 / 2
                exls_sheet["D24"] = "мес"

                capitalization_rate /= 100
                exls_sheet["B25"] = "Ставка капитализации Kp"
                exls_sheet["C25"] = str(round(capitalization_rate * 100, 1)) + "%"

                exls_sheet["B26"] = "Дата расчета Terminal value"
                exls_sheet["C26"] = Q_years_list[-1]

                rent_indexing /= 100
                exls_sheet["B27"] = "Ежегодная индексация аренды, %"
                exls_sheet["C27"] = str(round(rent_indexing * 100, 1)) + "%"

                minimal_ipc /= 100
                exls_sheet["B28"] = "Минимальный ИПЦ"

                exls_sheet["B29"] = "Средняя удельная КС недвижимости"
                exls_sheet["C29"] = self.division_into_categories(average_value_real_estate)
                exls_sheet["D29"] = "руб./кв.м."

                exls_sheet["B30"] = "НДС для УСН"
                exls_sheet["C30"] = str(nds_rate) + "%"

                exls_sheet["B31"] = "Дата оплаты первого налога на недвижимость"
                exls_sheet["C31"] = Q_years_list[Q_years_list.index(project_readiness) + 4]
                # exls_sheet["C31"] = Q_years_list[0]

                list_f_border_0 = ["B2", "B6", "B12", "C12", "B17", "B49", "B65"]
                for cell in list_f_border_0:
                    exls_sheet[cell].border = bordr_left_top

                list_f_border_1 = ["B3", "B9", "B14", "C14", "B30", "B58", "B74"]
                for cell in list_f_border_1:
                    exls_sheet[cell].border = bordr_left_bottom

                list_f_border_2 = ["C6", "D6", "D17", "C65"]
                for cell in list_f_border_2:
                    exls_sheet[cell].border = bordr_right_top

                list_f_border_3 = ["C9", "D9", "D30", "C74"]
                for cell in list_f_border_3:
                    exls_sheet[cell].border = bordr_right_bottom

                for r in range(50, 60):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(67, 76):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(62, 64):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(78, 80):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(8, 10):
                    for c in range(1, 5):
                        if c != 2:
                            exls_sheet.cell(row=r, column=c).border = bordr_right

                for r in range(14, 15):
                    for c in range(1, 3):
                        exls_sheet.cell(row=r, column=c).border = bordr_right

                for r in range(19, 31):
                    for c in range(1, 2):
                        exls_sheet.cell(row=r, column=c).border = bordr_right

                for r in range(19, 31):
                    for c in range(4, 5):
                        exls_sheet.cell(row=r, column=c).border = bordr_right

                exls_sheet["C17"].border = bordr_top
                exls_sheet["C31"].border = bordr_bottom

                exls_sheet["C2"].border = bordr_all
                exls_sheet["C3"].border = bordr_all

                lot_price = square * price_per_meter
                rent_qurtals_dict = {}
                rent_flow = 0
                for quartal in Q_years_list:
                    rent_qurtals_dict[quartal] = {
                        "auxiliary_calculation": 0.0,
                        "buying_property": 0.0,
                        "rent_flow": 0.0,
                        "income_tax": 0.0,
                        "renovation": 0.0,
                        "terminal_value": 0.0,
                        "property_tax": 0.0,
                        "money_flow": 0.0,
                    }
                    if quartal[0] == rental_date:
                        rent_flow += rent_flow * rent_indexing
                        rent_qurtals_dict[quartal]["auxiliary_calculation"] = rent_indexing
                    if quartal == renovation_quartal:
                        rent_qurtals_dict[quartal]["renovation"] = -square * price_of_finishing
                    if quartal == Q_years_list[Q_years_list.index(project_readiness) + rental_holidays - 1]:
                        rent_flow = rental_rate * square / 4
                    rent_qurtals_dict[quartal]["rent_flow"] = rent_flow
                    if quartal in transaction_dict.keys():
                        rent_qurtals_dict[quartal]["buying_property"] = -lot_price * transaction_dict[quartal]
                    if Q_years_list.index(quartal) == len(Q_years_list) - 1:
                        rent_qurtals_dict[quartal]["terminal_value"] = (
                                rent_qurtals_dict[quartal]["rent_flow"] * 4 / capitalization_rate
                        )
                    if Q_years_list.index(quartal) >= Q_years_list.index(project_readiness) + 4:
                        rent_qurtals_dict[quartal]["property_tax"] = -average_value_real_estate * square / 4 * 0.015

                    rent_qurtals_dict[quartal]["income_tax"] = (
                        -((rent_qurtals_dict[quartal]["rent_flow"] + rent_qurtals_dict[quartal]["terminal_value"])
                          * (self.income_tax_share + (nds_rate / 100)))
                    )
                    rent_qurtals_dict[quartal]["money_flow"] = (
                            rent_qurtals_dict[quartal]["buying_property"]
                            + rent_qurtals_dict[quartal]["rent_flow"]
                            + rent_qurtals_dict[quartal]["income_tax"]
                            + rent_qurtals_dict[quartal]["renovation"]
                            + rent_qurtals_dict[quartal]["terminal_value"]
                            + rent_qurtals_dict[quartal]["property_tax"]
                    )

                rent_irr = (
                                   (1 + np_f.irr(
                                       [rent_qurtals_dict[key]["money_flow"] for key in
                                        rent_qurtals_dict.keys()])) ** 4 - 1
                           ) * 100
                if np.isnan(rent_irr):
                    rent_irr = -1
                ali_1 = Alignment(horizontal="center", vertical="center")
                for r, quartal in enumerate(Q_years_list):
                    col = get_column_letter(r + 4)
                    l_col = get_column_letter(r + 3)

                    exls_sheet.cell(row=66, column=r + 4).value = quartal
                    if str(rental_date) + "Q" in quartal:
                        exls_sheet.cell(row=65, column=r + 4).value = rent_indexing
                    else:
                        exls_sheet.cell(row=65, column=r + 4).value = 0

                    if quartal in transaction_dict.keys():
                        exls_sheet.cell(row=67, column=r + 4).value = str(transaction_dict[quartal] * 100) + "%"
                        exls_sheet.cell(row=68, column=r + 4).value = self.division_into_categories(
                            -transaction_dict[quartal] * square * price_per_meter
                        )
                        exls_sheet.cell(row=68, column=r + 4).font = ft_3
                    if rent_qurtals_dict[quartal]["rent_flow"] < 0:
                        exls_sheet.cell(row=69, column=r + 4).font = ft_3
                    exls_sheet.cell(row=69, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["rent_flow"]
                    )

                    if rent_qurtals_dict[quartal]["renovation"] < 0:
                        exls_sheet.cell(row=71, column=r + 4).font = ft_3
                    exls_sheet.cell(row=71, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["renovation"]
                    )

                    if rent_qurtals_dict[quartal]["terminal_value"] < 0:
                        exls_sheet.cell(row=72, column=r + 4).font = ft_3
                    exls_sheet.cell(row=72, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["terminal_value"]
                    )

                    if rent_qurtals_dict[quartal]["income_tax"] < 0:
                        exls_sheet.cell(row=70, column=r + 4).font = ft_3
                    exls_sheet.cell(row=70, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["income_tax"]
                    )

                    if rent_qurtals_dict[quartal]["property_tax"] < 0:
                        exls_sheet.cell(row=73, column=r + 4).font = ft_3
                    exls_sheet.cell(row=73, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["property_tax"]
                    )

                    if rent_qurtals_dict[quartal]["money_flow"] < 0:
                        exls_sheet.cell(row=74, column=r + 4).font = ft_3
                    exls_sheet.cell(row=74, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["money_flow"]
                    )

                    exls_sheet.cell(row=12, column=r + 4).fill = ff_3
                    exls_sheet.cell(row=13, column=r + 4).fill = ff_3
                    exls_sheet.cell(row=67, column=r + 4).fill = ff_4

                    for i in range(65, 75):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all

                    for i in range(12, 15):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all

                    exls_sheet.column_dimensions[col].width = 15

                for r in range(1, 150):
                    for c in range(3, 150):
                        exls_sheet.cell(row=r, column=c).alignment = ali_1
                year_dict = {}
                for r, year in enumerate(years_list):
                    col = get_column_letter(r + 4)
                    year = str(year)
                    year_dict[year] = {
                        "transactions_percent": sum(
                            [transaction_dict[quartal] for quartal in transaction_dict.keys() if year in quartal]
                        ),
                        "buying_property": sum(
                            [
                                rent_qurtals_dict[quartal]["buying_property"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "rent_flow": sum(
                            [rent_qurtals_dict[quartal]["rent_flow"] for quartal in rent_qurtals_dict.keys() if
                             year in quartal]
                        ),
                        "income_tax": sum(
                            [
                                rent_qurtals_dict[quartal]["income_tax"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "renovation": sum(
                            [
                                rent_qurtals_dict[quartal]["renovation"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "terminal_value": sum(
                            [
                                rent_qurtals_dict[quartal]["terminal_value"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "property_tax": sum(
                            [
                                rent_qurtals_dict[quartal]["property_tax"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "money_flow": sum(
                            [
                                rent_qurtals_dict[quartal]["money_flow"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                    }
                    exls_sheet.cell(row=49, column=r + 4).value = year

                    if year_dict[year]["transactions_percent"] < 0:
                        exls_sheet.cell(row=51, column=r + 4).font = ft_3
                    exls_sheet.cell(row=51, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["transactions_percent"]
                    )

                    if year_dict[year]["buying_property"] < 0:
                        exls_sheet.cell(row=52, column=r + 4).font = ft_3
                    exls_sheet.cell(row=52, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["buying_property"]
                    )

                    if year_dict[year]["rent_flow"] < 0:
                        exls_sheet.cell(row=53, column=r + 4).font = ft_3
                    exls_sheet.cell(row=53, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["rent_flow"])

                    if year_dict[year]["terminal_value"] < 0:
                        exls_sheet.cell(row=54, column=r + 4).font = ft_3
                    exls_sheet.cell(row=54, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["terminal_value"]
                    )

                    if year_dict[year]["renovation"] < 0:
                        exls_sheet.cell(row=55, column=r + 4).font = ft_3
                    exls_sheet.cell(row=55, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["renovation"])

                    if year_dict[year]["income_tax"] < 0:
                        exls_sheet.cell(row=56, column=r + 4).font = ft_3
                    exls_sheet.cell(row=56, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["income_tax"])

                    if year_dict[year]["property_tax"] < 0:
                        exls_sheet.cell(row=57, column=r + 4).font = ft_3
                    exls_sheet.cell(row=57, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["property_tax"])

                    if year_dict[year]["money_flow"] < 0:
                        exls_sheet.cell(row=58, column=r + 4).font = ft_3
                    exls_sheet.cell(row=58, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["money_flow"])

                    exls_sheet.cell(row=51, column=r + 4).fill = ff_4

                    for i in range(49, 59):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all

                exls_sheet["B48"] = "2. Сдача в аренду"
                exls_sheet["B51"] = "Покупка"
                exls_sheet["B52"] = "Покупка"
                exls_sheet["B53"] = "Арендный поток за весь период"
                exls_sheet["B54"] = "Terminal value "
                exls_sheet["B55"] = "Ремонт помещения (Fit-Out)"
                exls_sheet["B56"] = "Налог на доход УСН + НДС для УСН"
                exls_sheet["B57"] = "Налог на недвижимость"
                exls_sheet["B58"] = "Денежный поток"

                exls_sheet["B61"] = "Прибыль"
                exls_sheet["B62"] = "IRR (доходность за период 6 лет с даты покупки)"

                exls_sheet["B64"] = "Вспомогательный расчет"
                exls_sheet["B67"] = "Покупка"
                exls_sheet["B68"] = "Покупка"
                exls_sheet["B69"] = "Арендный поток за весь период"
                exls_sheet["B70"] = "Налог на доход УСН + НДС для УСН"
                exls_sheet["B71"] = "Ремонт помещения (Fit-Out)"
                exls_sheet["B72"] = "Terminal Value "
                exls_sheet["B73"] = "Налог на недвижимость"
                exls_sheet["B74"] = "Денежный поток"

                exls_sheet["B77"] = "Прибыль"
                exls_sheet["B78"] = "IRR (доходность за период 6 лет с даты покупки)"

                total_dict = {
                    "buying_property": sum(
                        [rent_qurtals_dict[quartal]["buying_property"] for quartal in rent_qurtals_dict.keys()]
                    ),
                    "rent_flow": sum([rent_qurtals_dict[quartal]["rent_flow"] for quartal in rent_qurtals_dict.keys()]),
                    "income_tax": sum(
                        [rent_qurtals_dict[quartal]["income_tax"] for quartal in rent_qurtals_dict.keys()]),
                    "renovation": sum(
                        [rent_qurtals_dict[quartal]["renovation"] for quartal in rent_qurtals_dict.keys()]),
                    "terminal_value": sum(
                        [rent_qurtals_dict[quartal]["terminal_value"] for quartal in rent_qurtals_dict.keys()]
                    ),
                    "property_tax": sum(
                        [rent_qurtals_dict[quartal]["property_tax"] for quartal in rent_qurtals_dict.keys()]),
                    "money_flow": sum(
                        [rent_qurtals_dict[quartal]["money_flow"] for quartal in rent_qurtals_dict.keys()]),
                }
                exls_sheet["C66"] = "Итого"

                exls_sheet["C67"] = "100%"

                if total_dict["buying_property"] < 0:
                    exls_sheet["C68"].font = ft_3
                    exls_sheet["C52"].font = ft_3
                exls_sheet["C68"] = self.division_into_categories(total_dict["buying_property"])
                exls_sheet["C52"] = self.division_into_categories(total_dict["buying_property"])

                if total_dict["rent_flow"] < 0:
                    exls_sheet["C69"].font = ft_3
                    exls_sheet["C53"].font = ft_3
                exls_sheet["C69"] = self.division_into_categories(total_dict["rent_flow"])
                exls_sheet["C53"] = self.division_into_categories(total_dict["rent_flow"])

                if total_dict["income_tax"] < 0:
                    exls_sheet["C70"].font = ft_3
                    exls_sheet["C56"].font = ft_3
                exls_sheet["C70"] = self.division_into_categories(total_dict["income_tax"])
                exls_sheet["C56"] = self.division_into_categories(total_dict["income_tax"])

                if total_dict["renovation"] < 0:
                    exls_sheet["C71"].font = ft_3
                    exls_sheet["C55"].font = ft_3
                exls_sheet["C71"] = self.division_into_categories(total_dict["renovation"])
                exls_sheet["C55"] = self.division_into_categories(total_dict["renovation"])

                if total_dict["terminal_value"] < 0:
                    exls_sheet["C72"].font = ft_3
                    exls_sheet["C54"].font = ft_3
                exls_sheet["C72"] = self.division_into_categories(total_dict["terminal_value"])
                exls_sheet["C54"] = self.division_into_categories(total_dict["terminal_value"])

                if total_dict["property_tax"] < 0:
                    exls_sheet["C73"].font = ft_3
                    exls_sheet["C57"].font = ft_3
                exls_sheet["C73"] = self.division_into_categories(total_dict["property_tax"])
                exls_sheet["C57"] = self.division_into_categories(total_dict["property_tax"])
                sale_tax = total_dict["property_tax"]
                rent_tax = float(total_dict["property_tax"]) + float(total_dict["income_tax"])

                if total_dict["money_flow"] < 0:
                    exls_sheet["C74"].font = ft_3
                    exls_sheet["C58"].font = ft_3
                    exls_sheet["C77"].font = ft_3
                    exls_sheet["C61"].font = ft_3
                exls_sheet["C74"] = self.division_into_categories(total_dict["money_flow"])
                exls_sheet["C58"] = self.division_into_categories(total_dict["money_flow"])
                exls_sheet["C77"] = self.division_into_categories(total_dict["money_flow"])
                exls_sheet["C61"] = self.division_into_categories(total_dict["money_flow"])
                rent_income = total_dict["money_flow"]
                sale_income = total_dict["money_flow"]

                exls_sheet["C49"] = "Итого"
                exls_sheet["C51"] = "100%"

                exls_sheet["C78"] = str(round(rent_irr, 1)) + "%"

                exls_sheet["C62"] = str(round(rent_irr, 1)) + "%"
                ft_1 = Font(name="Calibri", size=11)
                ft_11 = Font(bold=True, name="Calibri", size=14)
                exls_sheet["C78"].font = ft_11
                exls_sheet["C62"].font = ft_11
                ft_1 = Font(name="Calibri", size=11)
                ff_1 = PatternFill("solid", fgColor="0099CCFF")

                ff_2 = PatternFill("solid", fgColor="0099CCFF")
                ft_3 = Font()
                # применяем стиль к ячейкам
                exls_sheet["B1"].font = ft_1
                exls_sheet["B1"].fill = ff_1
                exls_sheet["B48"].font = ft_1
                exls_sheet["B48"].fill = ff_1

                exls_sheet["C20"].fill = ff_2
                exls_sheet["C21"].fill = ff_2
                exls_sheet["C29"].fill = ff_2
                exls_sheet["B51"].fill = ff_4
                exls_sheet["C51"].fill = ff_4
                exls_sheet["B67"].fill = ff_4
                exls_sheet["C67"].fill = ff_4

                exls_sheet["B61"].fill = ff_4
                exls_sheet["C61"].fill = ff_4
                exls_sheet["B62"].fill = ff_4
                exls_sheet["C62"].fill = ff_4

                exls_sheet["B77"].fill = ff_4
                exls_sheet["C77"].fill = ff_4
                exls_sheet["B78"].fill = ff_4
                exls_sheet["C78"].fill = ff_4

                for row in range(32, 47):
                    exls_sheet.row_dimensions[row].hidden = True

                # for key in qurtals_dict.keys():
                #    print(key, end=' ')
                #    for und_key in qurtals_dict[key]:
                #        print(f'{und_key[0:5]}: {qurtals_dict[key][und_key]}', end='\t')
                #    print()

                if create_xlsx:
                    table_buffer = io.BytesIO()
                    wb.save(table_buffer)
                    table_buffer.seek(0)

                    return table_buffer
                sale_property = 0
                sale_irr = 0
                added_value = 0
                result_dict = {
                    "buying_property": round(float(total_dict["buying_property"]) / 1000000, 1),
                    "sale_property": round(sale_property / 1000000, 1),
                    "sale_tax": round(sale_tax / 1000000, 1),
                    "rent_tax": round(rent_tax / 1000000, 1),
                    "price_of_finishing": round(float(total_dict["renovation"]) / 1000000, 1),
                    "rent_flow": round(float(total_dict["rent_flow"]) / 1000000, 1),
                    "terminal_value": round(float(total_dict["terminal_value"]) / 1000000, 1),
                    "sale_income": round(float(sale_income) / 1000000, 1),
                    "rent_income": round(float(rent_income) / 1000000, 1),
                    "added_value": round(float(added_value) / 1000000, 1),
                    "rent_irr": round(rent_irr, 1),
                    "sale_irr": round(sale_irr, 1),
                }

                span.set_status(Status(StatusCode.OK))
                return result_dict
            except Exception as err:
                span.record_exception(err)
                span.set_status(Status(StatusCode.ERROR, str(err)))
                raise err

    async def calc_finance_model_building_office(
            self,
            square: float,
            price_per_meter: float,
            distance_to_metro: float,
            metro_station_name: str,
            estate_category: str,
            project_readiness: str,
            nds_rate: int,
            transaction_dict: dict,
            create_xlsx: bool = False,
    ) -> dict | io.BytesIO:
        """
        Generate a irr calculations on the profitability of sale real estate.
        And generate excell sheet with calculations.

        Args:
        square (float) - Total area in square meter
        price_per_meter (float) - Price in rubles per square meter
        distance_to_metro (float) - The path from real estate to the subway in meter
        metro_station_name (str) - Name of the metro station
        estate_category (str) - Real estate category (A or B)
        number_of_quartals (int) - The number of quartals to calculate
        project_readiness (str) - quarter of the year of delivery of real estate in format (1Q2024)
        transaction_dict (dict) - information about the frequency and percentage of transactions
        average_value_real_estate (int) - Average specific value of real estate
        minimal_ipc (float) - minimal ipc
        create_report (bool) - Do I need create report in excel file
        rental_holidays (int) - Rental holidays (office renovation) in month
        price_of_finishing (float) - The price of finishing the room in rubles per square meter
        capitalization_rate (float) - Capitalization rate Kp in %
        rent_indexing (float) - Annual rental indexation in %

        Returns:
        exls_sheet:  Returns the excell sheet with calculations
        irr (float): Returns the rental price of real estate.
        """
        with self.tracer.start_as_current_span(
                "EstateCalculator.calc_finance_model_building_office",
                kind=SpanKind.INTERNAL,
                attributes={
                    "square": square,
                    "price_per_meter": price_per_meter,
                    "distance_to_metro": distance_to_metro,
                    "metro_station_name": metro_station_name,
                    "estate_category": estate_category,
                    "project_readiness": project_readiness,
                    "nds_rate": nds_rate,
                    "transaction_dict": str(transaction_dict),
                    "create_xlsx": create_xlsx
                }
        ) as span:
            try:
                minimal_ipc: float = 6.0
                number_of_quartals: int = 24
                rental_holidays: int = 6
                price_of_finishing: float = 70_000.0
                capitalization_rate: float = 8.5
                rent_indexing: float = 7.0

                # styles
                ff_3 = PatternFill("solid", fgColor="00FFFF99")
                ff_4 = PatternFill("solid", fgColor="00C0C0C0")
                thins = Side(border_style="thin", color="00808080")
                bordr_all = Border(left=thins, right=thins, bottom=thins, top=thins)
                bordr_top = Border(top=thins)
                bordr_bottom = Border(bottom=thins)
                bordr_left = Border(left=thins)
                bordr_right = Border(right=thins)
                bordr_right_top = Border(right=thins, top=thins)
                bordr_left_top = Border(left=thins, top=thins)
                bordr_right_bottom = Border(right=thins, bottom=thins)
                bordr_left_bottom = Border(left=thins, bottom=thins)

                os.makedirs("/src/reports", exist_ok=True)
                report_file_name = f"/src/reports/{uuid4()}.xlsx"
                wb = Workbook()

                exls_sheet = wb.active
                exls_sheet.title = "Стройка"

                exls_sheet.column_dimensions["B"].width = 70
                exls_sheet.column_dimensions["C"].width = 30
                exls_sheet[
                    "B1"] = "Финансовая модель оценки инвестиционной привлекательности офиса на стадии строительства"
                exls_sheet.merge_cells("B1:C1")

                exls_sheet["B2"] = "Дата сделки"
                exls_sheet["C2"] = self.first_quartal()

                exls_sheet["B3"] = "Строительная готовность"
                exls_sheet["C3"] = project_readiness

                exls_sheet["B5"] = "Данные по выбранному помещению"

                exls_sheet["B6"] = "Лот"
                exls_sheet["C6"] = "-------"

                exls_sheet["B7"] = "Общая площадь"
                exls_sheet["C7"] = self.division_into_categories(square)
                exls_sheet["D7"] = "кв.м"

                exls_sheet["B8"] = "Цена"
                exls_sheet["C8"] = self.division_into_categories(price_per_meter)
                exls_sheet["D8"] = "руб./ кв. м."

                exls_sheet["B9"] = "Цена Лота (Total)"
                exls_sheet["C9"] = self.division_into_categories(square * price_per_meter)
                exls_sheet["D9"] = "руб."

                EPA = self.EPA_validation(self.first_quartal(), project_readiness, transaction_dict)

                price_rva = await self.calculate_price(
                    square=square,
                    distance_to_metro=distance_to_metro,
                    metro_station_name=metro_station_name,
                    estate_category=estate_category,
                    strategy="price",
                )

                rental_rate = await self.calculate_price(
                    square=square,
                    distance_to_metro=distance_to_metro,
                    metro_station_name=metro_station_name,
                    estate_category=estate_category,
                    strategy="rent",
                )

                average_value_real_estate = await self.get_cadastral_value(metro_station_name=metro_station_name)

                Q_years_list, years_list = self.create_datas_list(EPA, project_readiness, number_of_quartals)
                sale_quartal = Q_years_list[Q_years_list.index(project_readiness) + 1]
                self.validation_transactions_dict(transaction_dict)
                transaction_dict = self.create_transactions_dict(transaction_dict)

                exls_sheet["B11"] = "Условия оплаты"

                exls_sheet["B12"] = "Размер"

                exls_sheet["B13"] = "Квартал"

                exls_sheet["B14"] = "Сумма траншей"
                ft_3 = Font(color="00FF0000")
                ft_gray = Font(color="FF6D6D6D", italic=True)

                for r, quartal in enumerate(Q_years_list):
                    exls_sheet.cell(row=13, column=r + 4).value = quartal
                    if quartal in transaction_dict.keys():
                        exls_sheet.cell(row=12, column=r + 4).value = str(transaction_dict[quartal] * 100) + "%"
                        exls_sheet.cell(row=14, column=r + 4).value = self.division_into_categories(
                            -transaction_dict[quartal] * square * price_per_meter
                        )
                        exls_sheet.cell(row=14, column=r + 4).font = ft_3
                exls_sheet["C14"] = self.division_into_categories(
                    -sum([transaction_dict[quartal] * square * price_per_meter for quartal in transaction_dict.keys()])
                )
                exls_sheet["C14"].font = ft_3
                exls_sheet["C12"] = "100%"

                exls_sheet["B16"] = "Базовые предпосылки к расчету"

                exls_sheet["B17"] = "Перепродажа"
                exls_sheet["C17"] = sale_quartal
                rental_holidays = rental_holidays // 3

                if int(project_readiness[0]) + rental_holidays <= 4:
                    rental_date = str(int(project_readiness[0]) + rental_holidays)
                else:
                    rental_date = str(int(project_readiness[0]) + rental_holidays - 4)

                if int(project_readiness[0]) + 1 <= 4:
                    renovation_quartal = str(int(project_readiness[0]) + 1) + "Q" + str(int(project_readiness[2:]))
                else:
                    renovation_quartal = str(int(project_readiness[0]) + 1 - 4) + "Q" + str(
                        int(project_readiness[2:]) + 1)

                if int(project_readiness[0]) + 2 <= 4:
                    rental_flow_quartal = str(int(project_readiness[0]) + 2) + "Q" + str(int(project_readiness[2:]))
                else:
                    rental_flow_quartal = str(int(project_readiness[0]) + 2 - 4) + "Q" + str(
                        int(project_readiness[2:]) + 1)

                exls_sheet["B18"] = "Получения арендного потока"
                exls_sheet["C18"] = str(rental_date) + "Q"
                exls_sheet["C19"] = rental_flow_quartal

                exls_sheet["B20"] = "Цена объекта на стадии РВЭ"
                exls_sheet["C20"] = self.division_into_categories(price_rva)
                exls_sheet["D20"] = "руб./кв. м."

                exls_sheet["B21"] = "Ставка аренды по объекту на дату покупки"
                exls_sheet["C21"] = self.division_into_categories(rental_rate)
                exls_sheet["D21"] = "руб./кв.м./год"
                ipc_rent_rate = 0
                counter = None
                for r, quartal in enumerate(Q_years_list):
                    if type(counter) is float:
                        counter += 0.25
                        exls_sheet.cell(row=37, column=r + 4).value = counter
                        exls_sheet.cell(row=68, column=r + 4).value = counter
                    if quartal == Q_years_list[0]:
                        counter = 0.0
                        exls_sheet.cell(row=37, column=r + 4).value = counter
                        exls_sheet.cell(row=68, column=r + 4).value = counter

                    if quartal == project_readiness:
                        ipc_rent_rate = counter + 0.5
                        exls_sheet.cell(row=69, column=r + 4).value = ipc_rent_rate
                    else:
                        exls_sheet.cell(row=69, column=r + 4).value = 0

                predict_rent_rate = rental_rate * (1 + ipc_rent_rate * (minimal_ipc / 100))
                exls_sheet["B22"] = "Прогнозируемая ставка аренды объекта РВЭ с поправкой на ИПЦ"
                exls_sheet["C22"] = self.division_into_categories(predict_rent_rate)
                exls_sheet["D22"] = "руб./кв.м./год"

                exls_sheet["B23"] = "Месячная арендная плата (МАП)"
                m_a_p = rental_rate / 12 * square
                exls_sheet["C23"] = self.division_into_categories(m_a_p)
                m_a_p = rental_rate / 4 * square
                exls_sheet["D23"] = "руб. в месяц"

                exls_sheet["B24"] = "Цена отделки на дату покупки "
                exls_sheet["C24"] = self.division_into_categories(price_of_finishing)
                exls_sheet["D24"] = "руб./кв.м."

                predict_finishing_price = price_of_finishing * (1 + ipc_rent_rate * (minimal_ipc / 100))
                exls_sheet["B25"] = "Прогнозируемая цена отделки помещения (Fit-Out) на дату РвЭ"
                exls_sheet["C25"] = self.division_into_categories(predict_finishing_price)
                exls_sheet["D25"] = "руб./кв.м."

                exls_sheet["B26"] = "Арендные каникулы (ремонт офиса)"
                exls_sheet["C26"] = rental_holidays * 3
                exls_sheet["D26"] = "мес"

                capitalization_rate /= 100
                exls_sheet["B27"] = "Ставка капитализации Kp"
                exls_sheet["C27"] = str(round(capitalization_rate * 100, 1)) + "%"

                exls_sheet["B28"] = "Дата расчета Terminal value"
                exls_sheet["C28"] = Q_years_list[-1]

                rent_indexing /= 100
                exls_sheet["B29"] = "Ежегодная индексация аренды, %"
                exls_sheet["C29"] = str(round(rent_indexing * 100, 1)) + "%"

                minimal_ipc /= 100
                exls_sheet["B28"] = "Минимальный ИПЦ по росту цен и ставок аренды за период строительства"
                exls_sheet["C28"] = str(round(minimal_ipc * 100, 1)) + "%"

                exls_sheet["B29"] = "Средняя удельная КС недвижимости"
                exls_sheet["C29"] = self.division_into_categories(average_value_real_estate)
                exls_sheet["D29"] = "руб./кв.м."

                exls_sheet["B30"] = "Дата оплаты первого налога на недвижимость"
                exls_sheet["C30"] = Q_years_list[Q_years_list.index(project_readiness) + 4]

                exls_sheet["B31"] = "НДС для УСН"
                exls_sheet["C31"] = str(nds_rate) + "%"

                exls_sheet["B36"] = "Инвестиционные стратегии"
                exls_sheet["B37"] = "1. Покупка-Продажа"
                exls_sheet["C37"] = "Степень ИПЦ поквартально"

                exls_sheet["B39"] = "Покупка"
                exls_sheet["B40"] = "Покупка"
                exls_sheet["B41"] = "Продажа"
                exls_sheet["B42"] = "Добавленная стомость по ИПЦ"
                exls_sheet["B43"] = "Налог на доход УСН + НДС для УСН"
                exls_sheet["B44"] = "Денежный поток"
                sale_list = []
                counter = None

                list_f_border_0 = ["B2", "B6", "B12", "C12", "B17", "B43", "B49", "B65"]
                for cell in list_f_border_0:
                    exls_sheet[cell].border = bordr_left_top

                list_f_border_1 = ["B3", "B9", "B14", "C14", "B30", "B45", "B58", "B74"]
                for cell in list_f_border_1:
                    exls_sheet[cell].border = bordr_left_bottom

                list_f_border_2 = ["C6", "D6", "D17", "C65"]
                for cell in list_f_border_2:
                    exls_sheet[cell].border = bordr_right_top

                list_f_border_3 = ["C9", "D9", "D30", "C74"]
                for cell in list_f_border_3:
                    exls_sheet[cell].border = bordr_right_bottom

                for r in range(34, 42):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(53, 63):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(66, 75):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(63, 65):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(82, 84):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(46, 49):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(7, 9):
                    for c in range(1, 5):
                        if c != 2:
                            exls_sheet.cell(row=r, column=c).border = bordr_right

                for r in range(13, 14):
                    for c in range(1, 3):
                        exls_sheet.cell(row=r, column=c).border = bordr_right

                for r in range(18, 30):
                    for c in range(1, 2):
                        exls_sheet.cell(row=r, column=c).border = bordr_right

                for r in range(18, 30):
                    for c in range(4, 5):
                        exls_sheet.cell(row=r, column=c).border = bordr_right

                exls_sheet["C17"].border = bordr_top
                exls_sheet["C30"].border = bordr_bottom

                exls_sheet["B34"].border = bordr_all
                exls_sheet["C2"].border = bordr_all
                exls_sheet["C3"].border = bordr_all

                lot_price = square * price_per_meter

                sale_qurtals_dict = {}
                last_quartal = None
                for quartal in Q_years_list:
                    sale_qurtals_dict[quartal] = {
                        "auxiliary_calculation": 0.0,
                        "buying_property": 0.0,
                        "sale_property": 0.0,
                        "added_value": 0.0,
                        "property_tax": 0.0,
                        "money_flow": 0.0,
                    }
                    if last_quartal:
                        if (
                                sale_qurtals_dict[last_quartal]["auxiliary_calculation"] > 0
                                or sale_qurtals_dict[last_quartal]["buying_property"] < 0
                        ):
                            sale_qurtals_dict[quartal]["auxiliary_calculation"] = (
                                    sale_qurtals_dict[last_quartal]["auxiliary_calculation"] + 0.25
                            )

                    if quartal == sale_quartal:
                        sale_qurtals_dict[quartal]["sale_property"] = price_rva * square

                    if quartal in transaction_dict.keys():
                        sale_qurtals_dict[quartal]["buying_property"] = -lot_price * transaction_dict[quartal]

                    sale_qurtals_dict[quartal]["added_value"] = -sale_qurtals_dict[quartal]["sale_property"] + (
                            sale_qurtals_dict[quartal]["sale_property"]
                            * (1 + minimal_ipc) ** sale_qurtals_dict[quartal]["auxiliary_calculation"]
                    )
                    sale_qurtals_dict[quartal]["property_tax"] = (
                            -(sale_qurtals_dict[quartal]["sale_property"] + sale_qurtals_dict[quartal]["added_value"])
                            * (self.income_tax_share + (nds_rate / 100))
                    )
                    sale_qurtals_dict[quartal]["money_flow"] = (
                            sale_qurtals_dict[quartal]["buying_property"]
                            + sale_qurtals_dict[quartal]["sale_property"]
                            + sale_qurtals_dict[quartal]["added_value"]
                            + sale_qurtals_dict[quartal]["property_tax"]
                    )

                    last_quartal = quartal

                # for key in sale_qurtals_dict.keys():
                #    print(key, end=' ')
                #    for und_key in sale_qurtals_dict[key]:
                #        print(f'{und_key[0:5]}: {sale_qurtals_dict[key][und_key]}', end='\t')
                #    print()
                sale_irr = (
                                   (1 + np_f.irr(
                                       [sale_qurtals_dict[key]["money_flow"] for key in
                                        sale_qurtals_dict.keys()])) ** 4 - 1
                           ) * 100
                if not sale_irr:
                    sale_irr = -1

                exls_sheet["C48"] = str(round(sale_irr, 1)) + "%"

                rent_flow = 0
                rent_qurtals_dict = {}
                for quartal in Q_years_list:
                    rent_qurtals_dict[quartal] = {
                        "auxiliary_calculation": 0.0,
                        "buying_property": 0.0,
                        "rent_flow": 0.0,
                        "income_tax": 0.0,
                        "renovation": 0.0,
                        "terminal_value": 0.0,
                        "property_tax": 0.0,
                        "money_flow": 0.0,
                    }

                    if quartal[0] == rental_date:
                        rent_flow += rent_flow * rent_indexing
                        rent_qurtals_dict[quartal]["auxiliary_calculation"] = rent_indexing
                    if quartal == renovation_quartal:
                        rent_qurtals_dict[quartal]["renovation"] = -square * predict_finishing_price
                    if quartal == Q_years_list[Q_years_list.index(project_readiness) + rental_holidays]:
                        rent_flow = predict_rent_rate * square / 4
                    rent_qurtals_dict[quartal]["rent_flow"] = rent_flow
                    if quartal in transaction_dict.keys():
                        rent_qurtals_dict[quartal]["buying_property"] = -lot_price * transaction_dict[quartal]
                    if Q_years_list.index(quartal) == len(Q_years_list) - 1:
                        rent_qurtals_dict[quartal]["terminal_value"] = (
                                rent_qurtals_dict[quartal]["rent_flow"] * 4 / capitalization_rate
                        )
                    if Q_years_list.index(quartal) >= Q_years_list.index(project_readiness) + 4:
                        rent_qurtals_dict[quartal]["property_tax"] = -average_value_real_estate * square / 4 * 0.015

                    rent_qurtals_dict[quartal]["income_tax"] = (
                            -(rent_qurtals_dict[quartal]["rent_flow"] + rent_qurtals_dict[quartal]["terminal_value"])
                            * (self.income_tax_share + nds_rate / 100)
                    )
                    rent_qurtals_dict[quartal]["money_flow"] = (
                            rent_qurtals_dict[quartal]["buying_property"]
                            + rent_qurtals_dict[quartal]["rent_flow"]
                            + rent_qurtals_dict[quartal]["income_tax"]
                            + rent_qurtals_dict[quartal]["renovation"]
                            + rent_qurtals_dict[quartal]["terminal_value"]
                            + rent_qurtals_dict[quartal]["property_tax"]
                    )

                rent_irr = (
                                   (1 + np_f.irr(
                                       [rent_qurtals_dict[key]["money_flow"] for key in
                                        rent_qurtals_dict.keys()])) ** 4 - 1
                           ) * 100
                if np.isnan(rent_irr):
                    rent_irr = -1
                ali_1 = Alignment(horizontal="center", vertical="center")
                for r, quartal in enumerate(Q_years_list):
                    col = get_column_letter(r + 4)
                    l_col = get_column_letter(r + 3)

                    exls_sheet.cell(row=37, column=r + 4).font = ft_gray
                    exls_sheet.cell(row=68, column=r + 4).font = ft_gray
                    exls_sheet.cell(row=69, column=r + 4).font = ft_gray

                    exls_sheet.cell(row=38, column=r + 4).value = quartal
                    if quartal in transaction_dict.keys():
                        exls_sheet.cell(row=39, column=r + 4).value = str(transaction_dict[quartal] * 100) + "%"
                        exls_sheet.cell(row=40, column=r + 4).value = self.division_into_categories(
                            -transaction_dict[quartal] * square * price_per_meter
                        )
                        exls_sheet.cell(row=40, column=r + 4).font = ft_3

                    if sale_qurtals_dict[quartal]["sale_property"] < 0:
                        exls_sheet.cell(row=41, column=r + 4).font = ft_3
                    exls_sheet.cell(row=41, column=r + 4).value = self.division_into_categories(
                        sale_qurtals_dict[quartal]["sale_property"]
                    )

                    if sale_qurtals_dict[quartal]["added_value"] < 0:
                        exls_sheet.cell(row=42, column=r + 4).font = ft_3
                    exls_sheet.cell(row=42, column=r + 4).value = self.division_into_categories(
                        sale_qurtals_dict[quartal]["added_value"]
                    )

                    if sale_qurtals_dict[quartal]["property_tax"] < 0:
                        exls_sheet.cell(row=43, column=r + 4).font = ft_3
                    exls_sheet.cell(row=43, column=r + 4).value = self.division_into_categories(
                        sale_qurtals_dict[quartal]["property_tax"]
                    )

                    if sale_qurtals_dict[quartal]["money_flow"] < 0:
                        exls_sheet.cell(row=44, column=r + 4).font = ft_3
                    exls_sheet.cell(row=44, column=r + 4).value = self.division_into_categories(
                        sale_qurtals_dict[quartal]["money_flow"]
                    )

                    exls_sheet.cell(row=71, column=r + 4).value = quartal
                    if str(rental_date) + "Q" in quartal:
                        exls_sheet.cell(row=70, column=r + 4).value = (str(round((rent_indexing * 100), 1))) + "%"
                    else:
                        exls_sheet.cell(row=70, column=r + 4).value = 0

                    if quartal in transaction_dict.keys():
                        exls_sheet.cell(row=72, column=r + 4).value = str(transaction_dict[quartal] * 100) + "%"
                        exls_sheet.cell(row=73, column=r + 4).value = self.division_into_categories(
                            -transaction_dict[quartal] * square * price_per_meter
                        )
                        exls_sheet.cell(row=73, column=r + 4).font = ft_3

                    if rent_qurtals_dict[quartal]["rent_flow"] < 0:
                        exls_sheet.cell(row=74, column=r + 4).font = ft_3
                    exls_sheet.cell(row=74, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["rent_flow"]
                    )

                    if rent_qurtals_dict[quartal]["renovation"] < 0:
                        exls_sheet.cell(row=76, column=r + 4).font = ft_3
                    exls_sheet.cell(row=76, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["renovation"]
                    )

                    if rent_qurtals_dict[quartal]["terminal_value"] < 0:
                        exls_sheet.cell(row=77, column=r + 4).font = ft_3
                    exls_sheet.cell(row=77, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["terminal_value"]
                    )

                    if rent_qurtals_dict[quartal]["income_tax"] < 0:
                        exls_sheet.cell(row=75, column=r + 4).font = ft_3
                    exls_sheet.cell(row=75, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["income_tax"]
                    )

                    if rent_qurtals_dict[quartal]["property_tax"] < 0:
                        exls_sheet.cell(row=78, column=r + 4).font = ft_3
                    exls_sheet.cell(row=78, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["property_tax"]
                    )

                    if rent_qurtals_dict[quartal]["money_flow"] < 0:
                        exls_sheet.cell(row=79, column=r + 4).font = ft_3
                    exls_sheet.cell(row=79, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["money_flow"]
                    )

                    exls_sheet.cell(row=12, column=r + 4).fill = ff_3
                    exls_sheet.cell(row=13, column=r + 4).fill = ff_3
                    exls_sheet.cell(row=36, column=r + 4).fill = ff_4
                    exls_sheet.cell(row=67, column=r + 4).fill = ff_4
                    for i in range(35, 42):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all

                    for i in range(65, 75):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all

                    for i in range(12, 15):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all

                    exls_sheet.column_dimensions[col].width = 15

                for r in range(1, 150):
                    for c in range(3, 150):
                        exls_sheet.cell(row=r, column=c).alignment = ali_1
                year_dict = {}
                for r, year in enumerate(years_list):
                    col = get_column_letter(r + 4)
                    year = str(year)
                    year_dict[year] = {
                        "transactions_percent": sum(
                            [transaction_dict[quartal] for quartal in transaction_dict.keys() if year in quartal]
                        ),
                        "buying_property": sum(
                            [
                                rent_qurtals_dict[quartal]["buying_property"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "rent_flow": sum(
                            [rent_qurtals_dict[quartal]["rent_flow"] for quartal in rent_qurtals_dict.keys() if
                             year in quartal]
                        ),
                        "income_tax": sum(
                            [
                                rent_qurtals_dict[quartal]["income_tax"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "renovation": sum(
                            [
                                rent_qurtals_dict[quartal]["renovation"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "terminal_value": sum(
                            [
                                rent_qurtals_dict[quartal]["terminal_value"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "property_tax": sum(
                            [
                                rent_qurtals_dict[quartal]["property_tax"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "money_flow": sum(
                            [
                                rent_qurtals_dict[quartal]["money_flow"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                    }
                    exls_sheet.cell(row=52, column=r + 4).value = year

                    if year_dict[year]["transactions_percent"] < 0:
                        exls_sheet.cell(row=53, column=r + 4).font = ft_3
                    exls_sheet.cell(row=53, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["transactions_percent"]
                    )

                    if year_dict[year]["buying_property"] < 0:
                        exls_sheet.cell(row=54, column=r + 4).font = ft_3
                    exls_sheet.cell(row=54, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["buying_property"]
                    )

                    if year_dict[year]["rent_flow"] < 0:
                        exls_sheet.cell(row=55, column=r + 4).font = ft_3
                    exls_sheet.cell(row=55, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["rent_flow"])

                    if year_dict[year]["terminal_value"] < 0:
                        exls_sheet.cell(row=56, column=r + 4).font = ft_3
                    exls_sheet.cell(row=56, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["terminal_value"]
                    )

                    if year_dict[year]["renovation"] < 0:
                        exls_sheet.cell(row=57, column=r + 4).font = ft_3
                    exls_sheet.cell(row=57, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["renovation"])

                    if year_dict[year]["income_tax"] < 0:
                        exls_sheet.cell(row=58, column=r + 4).font = ft_3
                    exls_sheet.cell(row=58, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["income_tax"])

                    if year_dict[year]["property_tax"] < 0:
                        exls_sheet.cell(row=59, column=r + 4).font = ft_3
                    exls_sheet.cell(row=59, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["property_tax"])

                    if year_dict[year]["money_flow"] < 0:
                        exls_sheet.cell(row=60, column=r + 4).font = ft_3
                    exls_sheet.cell(row=60, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["money_flow"])

                    exls_sheet.cell(row=53, column=r + 4).fill = ff_4

                    for i in range(49, 59):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all
                total_dict = {
                    "buying_property": sum(
                        [sale_qurtals_dict[quartal]["buying_property"] for quartal in sale_qurtals_dict.keys()]
                    ),
                    "sale_property": sum(
                        [sale_qurtals_dict[quartal]["sale_property"] for quartal in sale_qurtals_dict.keys()]),
                    "added_value": sum(
                        [sale_qurtals_dict[quartal]["added_value"] for quartal in sale_qurtals_dict.keys()]),
                    "property_tax": sum(
                        [sale_qurtals_dict[quartal]["property_tax"] for quartal in sale_qurtals_dict.keys()]),
                    "money_flow": sum(
                        [sale_qurtals_dict[quartal]["money_flow"] for quartal in sale_qurtals_dict.keys()]),
                }
                exls_sheet["C38"] = "Итого"
                exls_sheet["C39"] = "100%"
                if total_dict["buying_property"] < 0:
                    exls_sheet["C40"].font = ft_3
                exls_sheet["C40"] = self.division_into_categories(total_dict["buying_property"])

                if total_dict["sale_property"] < 0:
                    exls_sheet["C41"].font = ft_3
                exls_sheet["C41"] = self.division_into_categories(total_dict["sale_property"])
                sale_property = total_dict["sale_property"]
                if total_dict["added_value"] < 0:
                    exls_sheet["C42"].font = ft_3
                exls_sheet["C42"] = self.division_into_categories(total_dict["added_value"])
                added_value = total_dict["added_value"]

                if total_dict["property_tax"] < 0:
                    exls_sheet["C43"].font = ft_3
                exls_sheet["C43"] = self.division_into_categories(total_dict["property_tax"])
                sale_tax = total_dict["property_tax"]

                if total_dict["money_flow"] < 0:
                    exls_sheet["C44"].font = ft_3
                    exls_sheet["C44"].font = ft_3
                exls_sheet["C44"] = self.division_into_categories(total_dict["money_flow"])
                sale_income = total_dict["money_flow"]

                exls_sheet["B46"] = "Прибыль"
                exls_sheet["C46"] = self.division_into_categories(total_dict["money_flow"])

                exls_sheet["B47"] = "Валовая доходность"
                exls_sheet["C47"] = str(round(total_dict["money_flow"] / lot_price * 100, 1)) + "%"

                exls_sheet["B48"] = "IRR"

                exls_sheet["B51"] = "2. Сдача в аренду"
                exls_sheet["B53"] = "Покупка"
                exls_sheet["B54"] = "Покупка"
                exls_sheet["B55"] = "Арендный поток за весь период"
                exls_sheet["B56"] = "Terminal value "
                exls_sheet["B57"] = "Ремонт помещения (Fit-Out)"
                exls_sheet["B58"] = "Налог на доход УСН 6% + НДС для УСН "
                exls_sheet["B59"] = "Налог на недвижимость"
                exls_sheet["B60"] = "Денежный поток"

                exls_sheet["B63"] = "Прибыль"
                exls_sheet["B64"] = "IRR (в периоде 6 лет с даты завершения строительства)"

                exls_sheet["B66"] = "Вспомогательный расчет"

                exls_sheet["B68"] = "Степень ИПЦ поквартально с даты покупки"
                exls_sheet["B69"] = "Степень ИПЦ для объекта на момент сдачи в аренду"
                exls_sheet["C69"] = str(ipc_rent_rate)
                exls_sheet["C69"].font = ft_gray

                exls_sheet["B72"] = "Покупка"
                exls_sheet["B73"] = "Покупка"
                exls_sheet["B74"] = "Арендный поток за весь период"
                exls_sheet["B75"] = "Налог на доход УСН 6% + НДС для УСН "
                exls_sheet["B76"] = "Ремонт помещения (Fit-Out)"
                exls_sheet["B77"] = "Terminal Value "
                exls_sheet["B78"] = "Налог на недвижимость"
                exls_sheet["B79"] = "Денежный поток"

                exls_sheet["B82"] = "Прибыль"
                exls_sheet["B83"] = "IRR (в периоде 6 лет с даты завершения строительства)"

                total_dict = {
                    "buying_property": sum(
                        [rent_qurtals_dict[quartal]["buying_property"] for quartal in rent_qurtals_dict.keys()]
                    ),
                    "rent_flow": sum([rent_qurtals_dict[quartal]["rent_flow"] for quartal in rent_qurtals_dict.keys()]),
                    "income_tax": sum(
                        [rent_qurtals_dict[quartal]["income_tax"] for quartal in rent_qurtals_dict.keys()]),
                    "renovation": sum(
                        [rent_qurtals_dict[quartal]["renovation"] for quartal in rent_qurtals_dict.keys()]),
                    "terminal_value": sum(
                        [rent_qurtals_dict[quartal]["terminal_value"] for quartal in rent_qurtals_dict.keys()]
                    ),
                    "property_tax": sum(
                        [rent_qurtals_dict[quartal]["property_tax"] for quartal in rent_qurtals_dict.keys()]),
                    "money_flow": sum(
                        [rent_qurtals_dict[quartal]["money_flow"] for quartal in rent_qurtals_dict.keys()]),
                }
                exls_sheet["C71"] = "Итого"

                exls_sheet["C72"] = "100%"

                if total_dict["buying_property"] < 0:
                    exls_sheet["C73"].font = ft_3
                    exls_sheet["C54"].font = ft_3
                exls_sheet["C73"] = self.division_into_categories(total_dict["buying_property"])
                exls_sheet["C54"] = self.division_into_categories(total_dict["buying_property"])

                if total_dict["rent_flow"] < 0:
                    exls_sheet["C74"].font = ft_3
                    exls_sheet["C55"].font = ft_3
                exls_sheet["C74"] = self.division_into_categories(total_dict["rent_flow"])
                exls_sheet["C55"] = self.division_into_categories(total_dict["rent_flow"])

                if total_dict["income_tax"] < 0:
                    exls_sheet["C75"].font = ft_3
                    exls_sheet["C58"].font = ft_3
                exls_sheet["C75"] = self.division_into_categories(total_dict["income_tax"])
                exls_sheet["C58"] = self.division_into_categories(total_dict["income_tax"])

                if total_dict["renovation"] < 0:
                    exls_sheet["C76"].font = ft_3
                    exls_sheet["C57"].font = ft_3
                exls_sheet["C76"] = self.division_into_categories(total_dict["renovation"])
                exls_sheet["C57"] = self.division_into_categories(total_dict["renovation"])

                if total_dict["terminal_value"] < 0:
                    exls_sheet["C77"].font = ft_3
                    exls_sheet["C56"].font = ft_3
                exls_sheet["C77"] = self.division_into_categories(total_dict["terminal_value"])
                exls_sheet["C56"] = self.division_into_categories(total_dict["terminal_value"])

                if total_dict["property_tax"] < 0:
                    exls_sheet["C78"].font = ft_3
                    exls_sheet["C59"].font = ft_3
                exls_sheet["C78"] = self.division_into_categories(total_dict["property_tax"])
                exls_sheet["C59"] = self.division_into_categories(total_dict["property_tax"])
                rent_tax = float(total_dict["property_tax"]) + float(total_dict["income_tax"])

                if total_dict["money_flow"] < 0:
                    exls_sheet["C79"].font = ft_3
                    exls_sheet["C60"].font = ft_3
                    exls_sheet["C82"].font = ft_3
                    exls_sheet["C63"].font = ft_3
                exls_sheet["C79"] = self.division_into_categories(total_dict["money_flow"])
                exls_sheet["C60"] = self.division_into_categories(total_dict["money_flow"])
                exls_sheet["C82"] = self.division_into_categories(total_dict["money_flow"])
                exls_sheet["C63"] = self.division_into_categories(total_dict["money_flow"])
                rent_income = total_dict["money_flow"]
                exls_sheet["C52"] = "Итого"
                exls_sheet["C53"] = "100%"

                exls_sheet["C83"] = str(round(rent_irr, 1)) + "%"

                exls_sheet["C64"] = str(round(rent_irr, 1)) + "%"

                ft_1 = Font(name="Calibri", size=11)
                ft_11 = Font(bold=True, name="Calibri", size=14)
                exls_sheet["C48"].font = ft_11
                exls_sheet["C83"].font = ft_11
                exls_sheet["C64"].font = ft_11
                ff_1 = PatternFill("solid", fgColor="0099CCFF")

                ff_2 = PatternFill("solid", fgColor="0099CCFF")

                # применяем стиль к ячейкам
                exls_sheet["B1"].font = ft_1
                exls_sheet["B1"].fill = ff_1

                exls_sheet["B51"].font = ft_1
                exls_sheet["B51"].fill = ff_1

                exls_sheet["C20"].fill = ff_2
                exls_sheet["C21"].fill = ff_2
                exls_sheet["C29"].fill = ff_2

                exls_sheet["B51"].fill = ff_4

                exls_sheet["B66"].fill = ff_4
                exls_sheet["C66"].fill = ff_4

                exls_sheet["B36"].fill = ff_4
                exls_sheet["C36"].fill = ff_4

                exls_sheet["B46"].fill = ff_4
                exls_sheet["C46"].fill = ff_4
                exls_sheet["B47"].fill = ff_4
                exls_sheet["C47"].fill = ff_4
                exls_sheet["B48"].fill = ff_4
                exls_sheet["C48"].fill = ff_4

                exls_sheet["B63"].fill = ff_4
                exls_sheet["C63"].fill = ff_4
                exls_sheet["B64"].fill = ff_4
                exls_sheet["C64"].fill = ff_4

                exls_sheet["B82"].fill = ff_4
                exls_sheet["C82"].fill = ff_4
                exls_sheet["B83"].fill = ff_4
                exls_sheet["C83"].fill = ff_4

                if create_xlsx:
                    table_buffer = io.BytesIO()
                    wb.save(table_buffer)
                    table_buffer.seek(0)

                    return table_buffer

                result_dict = {
                    "buying_property": round(float(total_dict["buying_property"]) / 1000000, 1),
                    "sale_property": round(sale_property / 1000000, 1),
                    "sale_tax": round(sale_tax / 1000000, 1),
                    "rent_tax": round(rent_tax / 1000000, 1),
                    "price_of_finishing": round(float(total_dict["renovation"]) / 1000000, 1),
                    "rent_flow": round(float(total_dict["rent_flow"]) / 1000000, 1),
                    "terminal_value": round(float(total_dict["terminal_value"]) / 1000000, 1),
                    "sale_income": round(float(sale_income) / 1000000, 1),
                    "rent_income": round(float(rent_income) / 1000000, 1),
                    "added_value": round(float(added_value) / 1000000, 1),
                    "rent_irr": round(rent_irr, 1),
                    "sale_irr": round(sale_irr, 1),
                }


                span.set_status(Status(StatusCode.OK))
                return result_dict
            except Exception as err:
                span.record_exception(err)
                span.set_status(Status(StatusCode.ERROR, str(err)))
                raise err

    async def calc_finance_model_building_retail(
            self,
            square: float,
            price_per_meter: float,
            transaction_dict: dict,
            price_rva: float,
            m_a_p: float,
            nds_rate: int,
            project_readiness: str,
            need_repairs: int,
            create_xlsx: bool = False,
    ) -> dict | io.BytesIO:
        """
        Generate a irr calculations on the profitability of sale retail.
        And generate excell sheet with calculations.

        Args:
        square (float) - Total area in square meter
        price_per_meter (float) - Price in rubles per square meter
        number_of_quartals (int) - The number of quartals to calculate
        project_readiness (str) - quarter of the year of delivery of real estate in format (1Q2024)
        transaction_dict (dict) - information about the frequency and percentage of transactions
        metro_station_name (str) - Name of the metro station
        price_rva (float) - Result cost of retail in rubles per square meter
        rental_rate (float) - Result rent of retail in rubles per square meter
        minimal_ipc (float) - minimal ipc
        create_report (bool) - Do I need create report in excel file
        rental_holidays (int) - Rental holidays (office renovation) in month
        price_of_finishing (float) - The price of finishing the room in rubles per square meter
        capitalization_rate (float) - Capitalization rate Kp in %
        rent_indexing (float) - Annual rental indexation in %

        Returns:
        exls_sheet:  Returns the excell sheet with calculations
        irr (float): Returns the rental price of real estate.
        """
        with self.tracer.start_as_current_span(
                "EstateCalculator.calc_finance_model_building_retail",
                kind=SpanKind.CLIENT,
                attributes={
                    "square": square,
                    "price_per_meter": price_per_meter,
                    "transaction_dict": str(transaction_dict),
                    "price_rva": price_rva,
                    "m_a_p": m_a_p,
                    "nds_rate": nds_rate,
                    "project_readiness": project_readiness,
                    "need_repairs": need_repairs,
                    "create_xlsx": create_xlsx
                }
        ) as span:
            try:
                minimal_ipc: float = 6.0
                number_of_quartals: int = 24
                rental_holidays: int = 6
                price_of_finishing: float = 70_000.0
                capitalization_rate: float = 8.5
                rent_indexing: float = 7.0

                if NeedRepairs.dont_need_repairs == need_repairs:
                    price_of_finishing = 0
                elif NeedRepairs.need_partial_repairs == need_repairs:
                    price_of_finishing *= 0.8

                # styles
                ff_3 = PatternFill("solid", fgColor="00FFFF99")
                ff_4 = PatternFill("solid", fgColor="00C0C0C0")
                thins = Side(border_style="thin", color="00808080")
                bordr_all = Border(left=thins, right=thins, bottom=thins, top=thins)
                bordr_top = Border(top=thins)
                bordr_bottom = Border(bottom=thins)
                bordr_left = Border(left=thins)
                bordr_right = Border(right=thins)
                bordr_right_top = Border(right=thins, top=thins)
                bordr_left_top = Border(left=thins, top=thins)
                bordr_right_bottom = Border(right=thins, bottom=thins)
                bordr_left_bottom = Border(left=thins, bottom=thins)

                os.makedirs("/src/reports", exist_ok=True)
                report_file_name = f"/src/reports/{uuid4()}.xlsx"
                wb = Workbook()

                exls_sheet = wb.active
                exls_sheet.title = "Стройка"

                exls_sheet.column_dimensions["B"].width = 70
                exls_sheet.column_dimensions["C"].width = 30
                exls_sheet[
                    "B1"] = "Финансовая модель оценки инвестиционной привлекательности объекта на стадии строительства"
                exls_sheet.merge_cells("B1:C1")

                exls_sheet["B2"] = "Дата сделки"
                exls_sheet["C2"] = self.first_quartal()

                exls_sheet["B3"] = "Строительная готовность"
                exls_sheet["C3"] = project_readiness

                exls_sheet["B5"] = "Данные по выбранному помещению"

                exls_sheet["B6"] = "Лот"
                exls_sheet["C6"] = "-------"

                exls_sheet["B7"] = "Общая площадь"
                exls_sheet["C7"] = square
                exls_sheet["D7"] = "кв.м"

                exls_sheet["B8"] = "Цена"
                exls_sheet["C8"] = self.division_into_categories(price_per_meter)
                exls_sheet["D8"] = "руб./ кв. м."

                exls_sheet["B9"] = "Цена Лота (Total)"
                exls_sheet["C9"] = self.division_into_categories(square * price_per_meter)
                exls_sheet["D9"] = "руб."

                EPA = self.EPA_validation(self.first_quartal(), project_readiness, transaction_dict)

                average_value_real_estate = 120000
                # average_value_real_estate = self.get_cadastral_value(metro_station_name=metro_station_name)

                Q_years_list, years_list = self.create_datas_list(EPA, project_readiness, number_of_quartals)
                sale_quartal = Q_years_list[Q_years_list.index(project_readiness)]
                self.validation_transactions_dict(transaction_dict)
                transaction_dict = self.create_transactions_dict(transaction_dict)

                exls_sheet["B11"] = "Условия рассрочки"

                exls_sheet["B12"] = "Размер"

                exls_sheet["B13"] = "Квартал"

                exls_sheet["B14"] = "Сумма траншей"
                ft_3 = Font(color="00FF0000")
                ft_gray = Font(color="FF6D6D6D", italic=True)

                for r, quartal in enumerate(Q_years_list):
                    exls_sheet.cell(row=13, column=r + 4).value = quartal
                    if quartal in transaction_dict.keys():
                        exls_sheet.cell(row=12, column=r + 4).value = str(
                            round(transaction_dict[quartal] * 100, 1)) + "%"
                        exls_sheet.cell(row=14, column=r + 4).value = self.division_into_categories(
                            -transaction_dict[quartal] * square * price_per_meter
                        )
                        exls_sheet.cell(row=14, column=r + 4).font = ft_3
                exls_sheet["C14"] = self.division_into_categories(
                    -sum([transaction_dict[quartal] * square * price_per_meter for quartal in transaction_dict.keys()])
                )
                exls_sheet["C14"].font = ft_3
                exls_sheet["C12"] = "100%"

                exls_sheet["B16"] = "Базовые предпосылки к расчету"

                exls_sheet["B17"] = "Перепродажа"
                exls_sheet["C17"] = sale_quartal
                rental_holidays = rental_holidays // 3
                if int(project_readiness[0]) + rental_holidays <= 4:
                    rental_date = str(int(project_readiness[0]) + rental_holidays)
                else:
                    rental_date = str(int(project_readiness[0]) + rental_holidays - 4)

                if int(project_readiness[0]) + 1 <= 4:
                    renovation_quartal = str(int(project_readiness[0]) + 1) + "Q" + str(int(project_readiness[2:]))
                else:
                    renovation_quartal = str(int(project_readiness[0]) + 1 - 4) + "Q" + str(
                        int(project_readiness[2:]) + 1)

                if int(project_readiness[0]) + 2 <= 4:
                    rental_flow_quartal = str(int(project_readiness[0]) + 2) + "Q" + str(int(project_readiness[2:]))
                else:
                    rental_flow_quartal = str(int(project_readiness[0]) + 2 - 4) + "Q" + str(
                        int(project_readiness[2:]) + 1)

                exls_sheet["B18"] = "Получения арендного потока"
                exls_sheet["C18"] = str(rental_date) + "Q"
                exls_sheet["C19"] = rental_flow_quartal

                exls_sheet["B20"] = "Цена на РВЭ"
                exls_sheet["C20"] = self.division_into_categories(price_rva)
                exls_sheet["D20"] = "руб./кв. м."

                rental_rate = m_a_p * 12 / square
                exls_sheet["B21"] = "Ставка аренды по объекту на дату покупки"
                exls_sheet["C21"] = self.division_into_categories(rental_rate)
                exls_sheet["D21"] = "руб./кв.м./год"

                ipc_rent_rate = 0
                counter = None
                for r, quartal in enumerate(Q_years_list):
                    if type(counter) is float:
                        counter += 0.25
                        exls_sheet.cell(row=37, column=r + 4).value = counter
                        exls_sheet.cell(row=68, column=r + 4).value = counter
                    if quartal == Q_years_list[0]:
                        counter = 0.0
                        exls_sheet.cell(row=37, column=r + 4).value = counter
                        exls_sheet.cell(row=68, column=r + 4).value = counter

                    if quartal == project_readiness:
                        ipc_rent_rate = counter + 0.5
                        exls_sheet.cell(row=69, column=r + 4).value = ipc_rent_rate
                    else:
                        exls_sheet.cell(row=69, column=r + 4).value = 0

                predict_rent_rate = rental_rate * (1 + ipc_rent_rate * (minimal_ipc / 100))
                exls_sheet["B22"] = "Прогнозируемая ставка аренды объекта РВЭ с поправкой на ИПЦ"
                exls_sheet["C22"] = self.division_into_categories(predict_rent_rate)
                exls_sheet["D22"] = "руб./кв.м./год"

                exls_sheet["B23"] = "Месячная арендная плата (МАП)"
                exls_sheet["C23"] = self.division_into_categories(m_a_p)
                exls_sheet["D23"] = "руб. в месяц"

                exls_sheet["B24"] = "Цена отделки на дату покупки "
                exls_sheet["C24"] = self.division_into_categories(price_of_finishing)
                exls_sheet["D24"] = "руб./кв.м."

                predict_finishing_price = price_of_finishing * (1 + ipc_rent_rate * (minimal_ipc / 100))
                exls_sheet["B25"] = "Прогнозируемая цена отделки помещения (Fit-Out) на дату РвЭ"
                exls_sheet["C25"] = self.division_into_categories(predict_finishing_price)
                exls_sheet["D25"] = "руб./кв.м."

                exls_sheet["B26"] = "Арендные каникулы (ремонт офиса)"
                exls_sheet["C26"] = rental_holidays * 3
                exls_sheet["D26"] = "мес"

                capitalization_rate /= 100
                exls_sheet["B27"] = "Ставка капитализации Kp"
                exls_sheet["C27"] = str(round(capitalization_rate * 100, 1)) + "%"

                exls_sheet["B28"] = "Дата расчета Terminal value"
                exls_sheet["C28"] = Q_years_list[-1]

                rent_indexing /= 100
                exls_sheet["B29"] = "Ежегодная индексация аренды, %"
                exls_sheet["C29"] = str(round(rent_indexing * 100, 1)) + "%"

                minimal_ipc /= 100
                exls_sheet["B30"] = "Минимальный ИПЦ по росту цен и ставок аренды за период строительства"
                exls_sheet["C30"] = str(round(minimal_ipc * 100, 1)) + "%"

                exls_sheet["B31"] = "НДС для УСН"
                exls_sheet["C31"] = str(nds_rate) + "%"

                exls_sheet["B32"] = "Средняя удельная КС недвижимости"
                exls_sheet["C32"] = self.division_into_categories(average_value_real_estate)
                exls_sheet["D32"] = "руб./кв.м."

                exls_sheet["B33"] = "Дата оплаты первого налога на недвижимость"
                exls_sheet["C33"] = Q_years_list[Q_years_list.index(project_readiness) + 3]

                exls_sheet["B36"] = "Инвестиционные стратегии"

                exls_sheet.row_dimensions[37].height = 30
                exls_sheet["B37"] = "1. Покупка-Продажа"
                exls_sheet["C37"] = "Степень ИПЦ поквартально"
                exls_sheet["C37"].font = ft_gray

                exls_sheet["B39"] = "Покупка"
                exls_sheet["B40"] = "Покупка"
                exls_sheet["B41"] = "Продажа"
                exls_sheet["B42"] = "Добавленная стомость по ИПЦ"
                exls_sheet["B43"] = "Налог на доход УСН + НДС для УСН"
                exls_sheet["B44"] = "Денежный поток"
                sale_list = []
                counter = None

                list_f_border_0 = ["B2", "B6", "B12", "C12", "B17", "B43", "B49", "B65"]
                for cell in list_f_border_0:
                    exls_sheet[cell].border = bordr_left_top

                list_f_border_1 = ["B3", "B9", "B14", "C14", "B30", "B45", "B58", "B74"]
                for cell in list_f_border_1:
                    exls_sheet[cell].border = bordr_left_bottom

                list_f_border_2 = ["C6", "D6", "D17", "C65"]
                for cell in list_f_border_2:
                    exls_sheet[cell].border = bordr_right_top

                list_f_border_3 = ["C9", "D9", "D30", "C74"]
                for cell in list_f_border_3:
                    exls_sheet[cell].border = bordr_right_bottom

                for r in range(34, 42):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(49, 59):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(66, 75):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(61, 63):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(77, 79):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(43, 46):
                    for c in range(2, 4):
                        exls_sheet.cell(row=r, column=c).border = bordr_all

                for r in range(7, 9):
                    for c in range(1, 5):
                        if c != 2:
                            exls_sheet.cell(row=r, column=c).border = bordr_right

                for r in range(13, 14):
                    for c in range(1, 3):
                        exls_sheet.cell(row=r, column=c).border = bordr_right

                for r in range(18, 30):
                    for c in range(1, 2):
                        exls_sheet.cell(row=r, column=c).border = bordr_right

                for r in range(18, 30):
                    for c in range(4, 5):
                        exls_sheet.cell(row=r, column=c).border = bordr_right

                exls_sheet["C17"].border = bordr_top
                exls_sheet["C30"].border = bordr_bottom

                exls_sheet["B34"].border = bordr_all
                exls_sheet["C2"].border = bordr_all
                exls_sheet["C3"].border = bordr_all

                lot_price = square * price_per_meter

                sale_qurtals_dict = {}
                last_quartal = None
                for quartal in Q_years_list:
                    sale_qurtals_dict[quartal] = {
                        "auxiliary_calculation": 0.0,
                        "buying_property": 0.0,
                        "sale_property": 0.0,
                        "added_value": 0.0,
                        "property_tax": 0.0,
                        "money_flow": 0.0,
                    }
                    if last_quartal:
                        if (
                                sale_qurtals_dict[last_quartal]["auxiliary_calculation"] > 0
                                or sale_qurtals_dict[last_quartal]["buying_property"] < 0
                        ):
                            sale_qurtals_dict[quartal]["auxiliary_calculation"] = (
                                    sale_qurtals_dict[last_quartal]["auxiliary_calculation"] + 0.25
                            )

                    if quartal == sale_quartal:
                        sale_qurtals_dict[quartal]["sale_property"] = price_rva * square

                    if quartal in transaction_dict.keys():
                        sale_qurtals_dict[quartal]["buying_property"] = -lot_price * transaction_dict[quartal]

                    sale_qurtals_dict[quartal]["added_value"] = -sale_qurtals_dict[quartal]["sale_property"] + (
                            sale_qurtals_dict[quartal]["sale_property"]
                            * (1 + minimal_ipc) ** sale_qurtals_dict[quartal]["auxiliary_calculation"]
                    )
                    sale_qurtals_dict[quartal]["property_tax"] = (
                            -(sale_qurtals_dict[quartal]["sale_property"] + sale_qurtals_dict[quartal]["added_value"])
                            * (self.income_tax_share + (nds_rate / 100))
                    )
                    sale_qurtals_dict[quartal]["money_flow"] = (
                            sale_qurtals_dict[quartal]["buying_property"]
                            + sale_qurtals_dict[quartal]["sale_property"]
                            + sale_qurtals_dict[quartal]["added_value"]
                            + sale_qurtals_dict[quartal]["property_tax"]
                    )

                    last_quartal = quartal

                # for key in sale_qurtals_dict.keys():
                #    print(key, end=' ')
                #    for und_key in sale_qurtals_dict[key]:
                #        print(f'{und_key[0:5]}: {sale_qurtals_dict[key][und_key]}', end='\t')
                #    print()
                sale_irr = (
                                   (1 + np_f.irr(
                                       [sale_qurtals_dict[key]["money_flow"] for key in
                                        sale_qurtals_dict.keys()])) ** 4 - 1
                           ) * 100
                if not sale_irr:
                    sale_irr = -1

                exls_sheet["C48"] = str(round(sale_irr, 1)) + "%"

                rent_flow = 0
                rent_qurtals_dict = {}
                for quartal in Q_years_list:
                    rent_qurtals_dict[quartal] = {
                        "auxiliary_calculation": 0.0,
                        "buying_property": 0.0,
                        "rent_flow": 0.0,
                        "income_tax": 0.0,
                        "renovation": 0.0,
                        "terminal_value": 0.0,
                        "property_tax": 0.0,
                        "money_flow": 0.0,
                    }
                    if quartal[0] == rental_date:
                        rent_flow += rent_flow * rent_indexing
                        rent_qurtals_dict[quartal]["auxiliary_calculation"] = rent_indexing
                    if quartal == renovation_quartal:
                        rent_qurtals_dict[quartal]["renovation"] = -square * predict_finishing_price
                    if quartal == Q_years_list[Q_years_list.index(project_readiness) + rental_holidays]:
                        rent_flow = predict_rent_rate * square / 4
                    rent_qurtals_dict[quartal]["rent_flow"] = rent_flow
                    if quartal in transaction_dict.keys():
                        rent_qurtals_dict[quartal]["buying_property"] = -lot_price * transaction_dict[quartal]

                    if Q_years_list.index(quartal) == len(Q_years_list) - 1:
                        rent_qurtals_dict[quartal]["terminal_value"] = (
                                rent_qurtals_dict[quartal]["rent_flow"] * 4 / capitalization_rate
                        )

                    if Q_years_list.index(quartal) >= Q_years_list.index(project_readiness) + 3:
                        rent_qurtals_dict[quartal]["property_tax"] = -average_value_real_estate * square / 4 * 0.015

                    rent_qurtals_dict[quartal]["income_tax"] = (
                            -(rent_qurtals_dict[quartal]["rent_flow"] + rent_qurtals_dict[quartal]["terminal_value"])
                            * (self.income_tax_share + nds_rate / 100)
                    )
                    rent_qurtals_dict[quartal]["money_flow"] = (
                            rent_qurtals_dict[quartal]["buying_property"]
                            + rent_qurtals_dict[quartal]["rent_flow"]
                            + rent_qurtals_dict[quartal]["income_tax"]
                            + rent_qurtals_dict[quartal]["renovation"]
                            + rent_qurtals_dict[quartal]["terminal_value"]
                            + rent_qurtals_dict[quartal]["property_tax"]
                    )

                rent_irr = (
                                   (1 + np_f.irr(
                                       [rent_qurtals_dict[key]["money_flow"] for key in
                                        rent_qurtals_dict.keys()])) ** 4 - 1
                           ) * 100
                if np.isnan(rent_irr):
                    rent_irr = -1

                ali_1 = Alignment(horizontal="center", vertical="center")
                for r, quartal in enumerate(Q_years_list):
                    col = get_column_letter(r + 4)
                    l_col = get_column_letter(r + 3)

                    exls_sheet.cell(row=37, column=r + 4).font = ft_gray
                    exls_sheet.cell(row=68, column=r + 4).font = ft_gray
                    exls_sheet.cell(row=69, column=r + 4).font = ft_gray

                    exls_sheet.cell(row=38, column=r + 4).value = quartal
                    if quartal in transaction_dict.keys():
                        exls_sheet.cell(row=39, column=r + 4).value = str(
                            round(transaction_dict[quartal] * 100, 1)) + "%"
                        exls_sheet.cell(row=40, column=r + 4).value = self.division_into_categories(
                            -transaction_dict[quartal] * square * price_per_meter
                        )
                        exls_sheet.cell(row=41, column=r + 4).font = ft_3

                    if sale_qurtals_dict[quartal]["sale_property"] < 0:
                        exls_sheet.cell(row=41, column=r + 4).font = ft_3
                    exls_sheet.cell(row=41, column=r + 4).value = self.division_into_categories(
                        sale_qurtals_dict[quartal]["sale_property"]
                    )

                    if sale_qurtals_dict[quartal]["added_value"] < 0:
                        exls_sheet.cell(row=42, column=r + 4).font = ft_3
                    exls_sheet.cell(row=42, column=r + 4).value = self.division_into_categories(
                        sale_qurtals_dict[quartal]["added_value"]
                    )

                    if sale_qurtals_dict[quartal]["property_tax"] < 0:
                        exls_sheet.cell(row=43, column=r + 4).font = ft_3
                    exls_sheet.cell(row=43, column=r + 4).value = self.division_into_categories(
                        sale_qurtals_dict[quartal]["property_tax"]
                    )

                    if sale_qurtals_dict[quartal]["money_flow"] < 0:
                        exls_sheet.cell(row=44, column=r + 4).font = ft_3
                    exls_sheet.cell(row=44, column=r + 4).value = self.division_into_categories(
                        sale_qurtals_dict[quartal]["money_flow"]
                    )

                    exls_sheet.cell(row=71, column=r + 4).value = quartal
                    if str(rental_date) + "Q" in quartal:
                        exls_sheet.cell(row=70, column=r + 4).value = rent_indexing
                    else:
                        exls_sheet.cell(row=70, column=r + 4).value = 0

                    if quartal in transaction_dict.keys():
                        exls_sheet.cell(row=72, column=r + 4).value = str(
                            round(transaction_dict[quartal] * 100, 1)) + "%"
                        exls_sheet.cell(row=73, column=r + 4).value = self.division_into_categories(
                            -transaction_dict[quartal] * square * price_per_meter
                        )
                        exls_sheet.cell(row=73, column=r + 4).font = ft_3

                    if rent_qurtals_dict[quartal]["rent_flow"] < 0:
                        exls_sheet.cell(row=74, column=r + 4).font = ft_3
                    exls_sheet.cell(row=74, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["rent_flow"]
                    )

                    if rent_qurtals_dict[quartal]["income_tax"] < 0:
                        exls_sheet.cell(row=75, column=r + 4).font = ft_3
                    exls_sheet.cell(row=75, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["income_tax"]
                    )

                    if rent_qurtals_dict[quartal]["renovation"] < 0:
                        exls_sheet.cell(row=76, column=r + 4).font = ft_3
                    exls_sheet.cell(row=76, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["renovation"]
                    )

                    if rent_qurtals_dict[quartal]["terminal_value"] < 0:
                        exls_sheet.cell(row=77, column=r + 4).font = ft_3
                    exls_sheet.cell(row=77, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["terminal_value"]
                    )

                    if rent_qurtals_dict[quartal]["property_tax"] < 0:
                        exls_sheet.cell(row=78, column=r + 4).font = ft_3
                    exls_sheet.cell(row=78, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["property_tax"]
                    )

                    if rent_qurtals_dict[quartal]["money_flow"] < 0:
                        exls_sheet.cell(row=79, column=r + 4).font = ft_3
                    exls_sheet.cell(row=79, column=r + 4).value = self.division_into_categories(
                        rent_qurtals_dict[quartal]["money_flow"]
                    )

                    exls_sheet.cell(row=13, column=r + 4).fill = ff_3
                    exls_sheet.cell(row=14, column=r + 4).fill = ff_3
                    # exls_sheet.cell(row=37, column=r + 4).fill = ff_4
                    # exls_sheet.cell(row=68, column=r + 4).fill = ff_4
                    for i in range(36, 43):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all

                    for i in range(66, 76):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all

                    for i in range(13, 16):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all

                    exls_sheet.column_dimensions[col].width = 15

                for r in range(1, 150):
                    for c in range(3, 150):
                        exls_sheet.cell(row=r, column=c).alignment = ali_1
                year_dict = {}
                for r, year in enumerate(years_list):
                    col = get_column_letter(r + 4)
                    year = str(year)
                    year_dict[year] = {
                        "transactions_percent": sum(
                            [transaction_dict[quartal] for quartal in transaction_dict.keys() if year in quartal]
                        ),
                        "buying_property": sum(
                            [
                                rent_qurtals_dict[quartal]["buying_property"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "rent_flow": sum(
                            [rent_qurtals_dict[quartal]["rent_flow"] for quartal in rent_qurtals_dict.keys() if
                             year in quartal]
                        ),
                        "income_tax": sum(
                            [
                                rent_qurtals_dict[quartal]["income_tax"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "renovation": sum(
                            [
                                rent_qurtals_dict[quartal]["renovation"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "terminal_value": sum(
                            [
                                rent_qurtals_dict[quartal]["terminal_value"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "property_tax": sum(
                            [
                                rent_qurtals_dict[quartal]["property_tax"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                        "money_flow": sum(
                            [
                                rent_qurtals_dict[quartal]["money_flow"]
                                for quartal in rent_qurtals_dict.keys()
                                if year in quartal
                            ]
                        ),
                    }
                    exls_sheet.cell(row=52, column=r + 4).value = year

                    if year_dict[year]["transactions_percent"] < 0:
                        exls_sheet.cell(row=53, column=r + 4).font = ft_3
                    exls_sheet.cell(row=53, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["transactions_percent"]
                    )

                    if year_dict[year]["buying_property"] < 0:
                        exls_sheet.cell(row=54, column=r + 4).font = ft_3
                    exls_sheet.cell(row=54, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["buying_property"]
                    )

                    if year_dict[year]["rent_flow"] < 0:
                        exls_sheet.cell(row=55, column=r + 4).font = ft_3
                    exls_sheet.cell(row=55, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["rent_flow"])

                    if year_dict[year]["terminal_value"] < 0:
                        exls_sheet.cell(row=56, column=r + 4).font = ft_3
                    exls_sheet.cell(row=56, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["terminal_value"]
                    )

                    if year_dict[year]["renovation"] < 0:
                        exls_sheet.cell(row=57, column=r + 4).font = ft_3
                    exls_sheet.cell(row=57, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["renovation"])

                    if year_dict[year]["income_tax"] < 0:
                        exls_sheet.cell(row=58, column=r + 4).font = ft_3
                    exls_sheet.cell(row=58, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["income_tax"])

                    if year_dict[year]["property_tax"] < 0:
                        exls_sheet.cell(row=59, column=r + 4).font = ft_3
                    exls_sheet.cell(row=59, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["property_tax"])

                    if year_dict[year]["money_flow"] < 0:
                        exls_sheet.cell(row=60, column=r + 4).font = ft_3
                    exls_sheet.cell(row=60, column=r + 4).value = self.division_into_categories(
                        year_dict[year]["money_flow"])

                    exls_sheet.cell(row=52, column=r + 4).fill = ff_4

                    for i in range(50, 60):
                        exls_sheet.cell(row=i, column=r + 4).border = bordr_all
                total_dict = {
                    "buying_property": sum(
                        [sale_qurtals_dict[quartal]["buying_property"] for quartal in sale_qurtals_dict.keys()]
                    ),
                    "sale_property": sum(
                        [sale_qurtals_dict[quartal]["sale_property"] for quartal in sale_qurtals_dict.keys()]),
                    "added_value": sum(
                        [sale_qurtals_dict[quartal]["added_value"] for quartal in sale_qurtals_dict.keys()]),
                    "property_tax": sum(
                        [sale_qurtals_dict[quartal]["property_tax"] for quartal in sale_qurtals_dict.keys()]),
                    "money_flow": sum(
                        [sale_qurtals_dict[quartal]["money_flow"] for quartal in sale_qurtals_dict.keys()]),
                }

                exls_sheet["C38"] = "Итого"
                exls_sheet["C39"] = "100%"
                if total_dict["buying_property"] < 0:
                    exls_sheet["C40"].font = ft_3
                exls_sheet["C40"] = self.division_into_categories(total_dict["buying_property"])

                if total_dict["sale_property"] < 0:
                    exls_sheet["C41"].font = ft_3
                exls_sheet["C41"] = self.division_into_categories(total_dict["sale_property"])
                sale_property = total_dict["sale_property"]

                if total_dict["added_value"] < 0:
                    exls_sheet["C42"].font = ft_3
                exls_sheet["C42"] = self.division_into_categories(total_dict["added_value"])
                added_value = total_dict["added_value"]

                if total_dict["property_tax"] < 0:
                    exls_sheet["C43"].font = ft_3
                exls_sheet["C43"] = self.division_into_categories(total_dict["property_tax"])
                sale_tax = total_dict["property_tax"]

                if total_dict["money_flow"] < 0:
                    exls_sheet["C44"].font = ft_3
                    exls_sheet["C46"].font = ft_3
                exls_sheet["C44"] = self.division_into_categories(total_dict["money_flow"])
                sale_income = total_dict["money_flow"]

                exls_sheet["B46"] = "Прибыль"
                exls_sheet["C46"] = self.division_into_categories(total_dict["money_flow"])

                exls_sheet["B47"] = "Валовая доходность"
                exls_sheet["C47"] = str(round(total_dict["money_flow"] / lot_price * 100, 1)) + "%"

                exls_sheet["B48"] = "IRR (доходность за период 6 лет с даты покупки)"

                exls_sheet["B51"] = "2. Сдача в аренду"
                exls_sheet["B53"] = "Покупка"
                exls_sheet["B54"] = "Покупка"
                exls_sheet["B55"] = "Арендный поток за весь период"
                exls_sheet["B56"] = "Terminal value "
                exls_sheet["B57"] = "Ремонт помещения (Fit-Out)"
                exls_sheet["B58"] = "Налог на доход УСН + НДС для УСН"
                exls_sheet["B59"] = "Налог на недвижимость"
                exls_sheet["B60"] = "Денежный поток"

                exls_sheet["B63"] = "Прибыль"
                exls_sheet["B64"] = "IRR (за период 6 лет с даты завершения строительства)"
                exls_sheet["C64"] = str(round(rent_irr, 1)) + "%"

                exls_sheet["B66"] = "Вспомогательный расчет"

                exls_sheet["B68"] = "Степень ИПЦ поквартально с даты покупки"
                exls_sheet["B69"] = "Степень ИПЦ для объекта на момент сдачи в аренду"
                exls_sheet["C69"] = str(ipc_rent_rate)
                exls_sheet["C69"].font = ft_gray

                exls_sheet["B72"] = "Покупка"
                exls_sheet["B73"] = "Покупка"
                exls_sheet["B74"] = "Арендный поток за весь период"
                exls_sheet["B75"] = "Налог на доход УСН + НДС для УСН"
                exls_sheet["B76"] = "Ремонт помещения (Fit-Out)"
                exls_sheet["B77"] = "Terminal Value "
                exls_sheet["B78"] = "Налог на недвижимость"
                exls_sheet["B79"] = "Денежный поток"

                exls_sheet["B82"] = "Прибыль"
                exls_sheet["B83"] = "IRR (за период 6 лет с даты завершения строительства)"
                exls_sheet["C83"] = str(round(rent_irr, 1)) + "%"

                total_dict = {
                    "buying_property": sum(
                        [rent_qurtals_dict[quartal]["buying_property"] for quartal in rent_qurtals_dict.keys()]
                    ),
                    "rent_flow": sum([rent_qurtals_dict[quartal]["rent_flow"] for quartal in rent_qurtals_dict.keys()]),
                    "income_tax": sum(
                        [rent_qurtals_dict[quartal]["income_tax"] for quartal in rent_qurtals_dict.keys()]),
                    "renovation": sum(
                        [rent_qurtals_dict[quartal]["renovation"] for quartal in rent_qurtals_dict.keys()]),
                    "terminal_value": sum(
                        [rent_qurtals_dict[quartal]["terminal_value"] for quartal in rent_qurtals_dict.keys()]
                    ),
                    "property_tax": sum(
                        [rent_qurtals_dict[quartal]["property_tax"] for quartal in rent_qurtals_dict.keys()]),
                    "money_flow": sum(
                        [rent_qurtals_dict[quartal]["money_flow"] for quartal in rent_qurtals_dict.keys()]),
                }
                exls_sheet["C71"] = "Итого"

                exls_sheet["C72"] = "100%"

                if total_dict["buying_property"] < 0:
                    exls_sheet["C73"].font = ft_3
                    exls_sheet["C54"].font = ft_3
                exls_sheet["C73"] = self.division_into_categories(total_dict["buying_property"])
                exls_sheet["C54"] = self.division_into_categories(total_dict["buying_property"])

                if total_dict["rent_flow"] < 0:
                    exls_sheet["C74"].font = ft_3
                    exls_sheet["C55"].font = ft_3
                exls_sheet["C74"] = self.division_into_categories(total_dict["rent_flow"])
                exls_sheet["C55"] = self.division_into_categories(total_dict["rent_flow"])

                if total_dict["income_tax"] < 0:
                    exls_sheet["C75"].font = ft_3
                    exls_sheet["C58"].font = ft_3
                exls_sheet["C75"] = self.division_into_categories(total_dict["income_tax"])
                exls_sheet["C58"] = self.division_into_categories(total_dict["income_tax"])

                if total_dict["renovation"] < 0:
                    exls_sheet["C76"].font = ft_3
                    exls_sheet["C57"].font = ft_3
                exls_sheet["C76"] = self.division_into_categories(total_dict["renovation"])
                exls_sheet["C57"] = self.division_into_categories(total_dict["renovation"])

                if total_dict["terminal_value"] < 0:
                    exls_sheet["C77"].font = ft_3
                    exls_sheet["C56"].font = ft_3
                exls_sheet["C77"] = self.division_into_categories(total_dict["terminal_value"])
                exls_sheet["C56"] = self.division_into_categories(total_dict["terminal_value"])

                if total_dict["property_tax"] < 0:
                    exls_sheet["C78"].font = ft_3
                    exls_sheet["C59"].font = ft_3
                exls_sheet["C78"] = self.division_into_categories(total_dict["property_tax"])
                exls_sheet["C59"] = self.division_into_categories(total_dict["property_tax"])
                rent_tax = float(total_dict["property_tax"]) + float(total_dict["income_tax"])

                if total_dict["money_flow"] < 0:
                    exls_sheet["C60"].font = ft_3
                    exls_sheet["C63"].font = ft_3
                    exls_sheet["C79"].font = ft_3
                    exls_sheet["C82"].font = ft_3
                exls_sheet["C60"] = self.division_into_categories(total_dict["money_flow"])
                exls_sheet["C63"] = self.division_into_categories(total_dict["money_flow"])
                exls_sheet["C79"] = self.division_into_categories(total_dict["money_flow"])
                exls_sheet["C82"] = self.division_into_categories(total_dict["money_flow"])
                rent_income = total_dict["money_flow"]

                exls_sheet["C52"] = "Итого"
                exls_sheet["C53"] = "100%"

                ft_1 = Font(name="Calibri", size=11)
                ft_11 = Font(bold=True, name="Calibri", size=14)
                exls_sheet["C48"].font = ft_11
                exls_sheet["C64"].font = ft_11
                exls_sheet["C83"].font = ft_11
                ff_1 = PatternFill("solid", fgColor="0099CCFF")

                ff_2 = PatternFill("solid", fgColor="0099CCFF")

                # применяем стиль к ячейкам
                exls_sheet["B1"].font = ft_1
                exls_sheet["B1"].fill = ff_1
                exls_sheet["B35"].font = ft_1
                exls_sheet["B35"].fill = ff_1
                exls_sheet["B49"].font = ft_1
                exls_sheet["B49"].fill = ff_1

                exls_sheet["C21"].fill = ff_2
                exls_sheet["C22"].fill = ff_2
                exls_sheet["C30"].fill = ff_2
                exls_sheet["B52"].fill = ff_4
                exls_sheet["C52"].fill = ff_4
                exls_sheet["B68"].fill = ff_4
                exls_sheet["C68"].fill = ff_4

                exls_sheet["B37"].fill = ff_4
                exls_sheet["C37"].fill = ff_4

                exls_sheet["B46"].fill = ff_4
                exls_sheet["C46"].fill = ff_4
                exls_sheet["B47"].fill = ff_4
                exls_sheet["C47"].fill = ff_4
                exls_sheet["B48"].fill = ff_4
                exls_sheet["C48"].fill = ff_4

                exls_sheet["B63"].fill = ff_4
                exls_sheet["C63"].fill = ff_4
                exls_sheet["B64"].fill = ff_4
                exls_sheet["C64"].fill = ff_4

                exls_sheet["B82"].fill = ff_4
                exls_sheet["C82"].fill = ff_4
                exls_sheet["B83"].fill = ff_4
                exls_sheet["C83"].fill = ff_4

                if create_xlsx:
                    table_buffer = io.BytesIO()
                    wb.save(table_buffer)
                    table_buffer.seek(0)

                    return table_buffer

                # is_ready =
                result_dict = {
                    "buying_property": round(float(total_dict["buying_property"]) / 1000000, 1),
                    "sale_property": round(sale_property / 1000000, 1),
                    "sale_tax": round(sale_tax / 1000000, 1),
                    "rent_tax": round(rent_tax / 1000000, 1),
                    "price_of_finishing": round(float(total_dict["renovation"]) / 1000000, 1),
                    "rent_flow": round(float(total_dict["rent_flow"]) / 1000000, 1),
                    "terminal_value": round(float(total_dict["terminal_value"]) / 1000000, 1),
                    "sale_income": round(float(sale_income) / 1000000, 1),
                    "rent_income": round(float(rent_income) / 1000000, 1),
                    "added_value": round(float(added_value) / 1000000, 1),
                    "rent_irr": round(rent_irr, 1),
                    "sale_irr": round(sale_irr, 1),
                }

                span.set_status(Status(StatusCode.OK))
                return result_dict
            except Exception as err:
                span.record_exception(err)
                span.set_status(Status(StatusCode.ERROR, str(err)))
                raise err

    def create_transactions_dict(self, transaction_dict):
        new_transaction_dict = {}
        for key in transaction_dict.keys():
            new_transaction_dict[key] = transaction_dict[key] / 100

        return new_transaction_dict

    def validation_transactions_dict(self, transaction_dict):
        transaction_sum = sum([transaction_dict[key] for key in transaction_dict.keys()])
        if not transaction_sum == 100:
            raise common.TransactionDictSumNotEqual100(
                f"Сумма транзакций равна: {transaction_sum}. Должна быть равна 100")

    def create_datas_list(self, EPA: str, project_readiness: str, number_of_quartals: int):
        quartal = 0
        Q_year = [int(param) for param in EPA.split("Q")]
        Q_years_list = [f"{str(Q_year[0])}Q{str(Q_year[1])}"]
        years_list = [Q_year[1]]
        readiness_flaq = False
        s = str(Q_year[0]) + "Q" + str(Q_year[1])

        while quartal != number_of_quartals:
            if s == project_readiness:
                readiness_flaq = True
            if readiness_flaq:
                quartal += 1
            if Q_year[0] % 4 == 0:
                Q_year[0] = 1
                Q_year[1] += 1
                years_list.append(str(Q_year[1]))
            else:
                Q_year[0] += 1
            s = str(Q_year[0]) + "Q" + str(Q_year[1])
            Q_years_list.append(s)

        return Q_years_list, years_list

    def EPA_validation(self, EPA, project_readiness, transaction_dict):
        Q_list = [EPA, project_readiness] + list(transaction_dict.keys())
        Q_split_list = []
        for quartal in Q_list:
            Q_split_list.append((quartal.split("Q")[0], quartal.split("Q")[1]))

        min_year = min([j[1] for j in Q_split_list])
        Q_split_list = [i for i in Q_split_list if i[1] == min_year]
        min_quartal = min([j[0] for j in Q_split_list])

        return str(min_quartal) + "Q" + min_year

    def first_quartal(self):
        date = DT.date.today()
        quartal = (date.month - 1) // 3 + 1

        return "{}Q{}".format(quartal, date.year)

    def division_into_categories(self, value):
        if value < 0:
            flaq = True
            value *= -1
        else:
            flaq = False
        value = str(int(value))
        if len(value) <= 3:
            if flaq:
                return "-" + value
            else:
                return value
        else:
            s = ""
            counter = 0
            for l in value[::-1]:
                counter += 1
                if counter != 3:
                    s += l
                else:
                    counter = 0
                    s += l + " "
            if len(value) % 3 == 0:
                if flaq:
                    return "-" + s[::-1][1:]
                else:
                    return s[::-1][1:]

            else:
                if flaq:
                    return "-" + s[::-1]
                else:
                    return s[::-1]

    async def calculate_price(
            self,
            square: float,
            distance_to_metro: float,
            metro_station_name: str,
            estate_category: str,
            strategy: str
    ) -> float:
        metro_station_name = metro_station_name.lower().replace("ё", "е")
        metro_coeffs = {}
        for metro_coeff in await self.metro_repo.all_metro_distance_coeff():
            metro_coeffs[(metro_coeff.min_distance, metro_coeff.max_distance)] = metro_coeff.coeff

        square_coeffs = {}
        for square_coeff in await self.metro_repo.all_square_coeff():
            square_coeffs[(square_coeff.min_square, square_coeff.max_square)] = square_coeff.coeff

        metro_station = await self.metro_repo.metro_station_by_name(metro_station_name)
        if not metro_station:
            raise common.MetroStationNotFound("Метро не найдено")

        metro_station = metro_station[0]

        way_coeff = self.create_coeff(metro_coeffs, distance_to_metro)
        square_coeff = self.create_coeff(square_coeffs, square)

        price_param = strategy + "_" + estate_category.lower()
        price = metro_station.to_dict()[price_param] * way_coeff * square_coeff

        return price

    async def get_cadastral_value(self, metro_station_name: str) -> float:
        metro_station_name = metro_station_name.lower().replace("ё", "е").rstrip()
        metro_station = await self.metro_repo.metro_station_by_name(metro_station_name)
        if not metro_station:
            raise common.MetroStationNotFound("Метро не найдено")

        metro_station = metro_station[0]

        return float(metro_station.average_cadastral_value)

    def create_coeff(self, coeffs_dict: dict, param: float) -> float:
        coefficient = None
        for key in coeffs_dict.keys():
            if float(key[0]) <= param <= float(key[1]):
                coefficient = float(coeffs_dict[key])
                break

        return coefficient
